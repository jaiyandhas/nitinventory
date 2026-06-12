"""Import parsing service for users and budgets."""
import io
import csv
import re
from datetime import datetime
from typing import Optional, Dict, Any

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import select, and_, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.user import User, Department, RoleManager
from app.models.budget import BudgetMaster, FinancialYear
from app.core.security import get_password_hash


class ImportService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def import_users(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """
        Parses an Excel file containing user onboarding details, resolves/creates departments/roles,
        and creates/updates user entities.
        """
        if not filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Only Excel files (.xlsx or .xls) are supported")

        try:
            wb = load_workbook(filename=io.BytesIO(file_content))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=400, detail="The sheet is empty")

        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        expected = ["name", "email", "department name", "role name"]
        
        # Check that headers match
        for exp in expected:
            if exp not in headers:
                raise HTTPException(status_code=400, detail=f"Missing column '{exp}' in header row")

        name_idx = headers.index("name")
        email_idx = headers.index("email")
        dept_idx = headers.index("department name")
        role_idx = headers.index("role-name") if "role-name" in headers else headers.index("role name")

        success_count = 0
        errors = []

        roles_cache = {}
        depts_cache = {}

        default_pw_hash = get_password_hash("Password@123")

        for row_num, row in enumerate(rows[1:], start=2):
            if not any(row):  # skip completely empty row
                continue
            
            name = str(row[name_idx]).strip() if row[name_idx] is not None else ""
            email = str(row[email_idx]).strip() if row[email_idx] is not None else ""
            dept_name = str(row[dept_idx]).strip() if row[dept_idx] is not None else ""
            role_name = str(row[role_idx]).strip() if row[role_idx] is not None else ""

            if not name or not email or not dept_name or not role_name:
                errors.append(f"Row {row_num}: Missing required field (all fields are required)")
                continue

            if "@" not in email:
                errors.append(f"Row {row_num}: Invalid email format '{email}'")
                continue

            try:
                # 1. Resolve department (case-insensitive)
                dept_key = dept_name.lower()
                if dept_key not in depts_cache:
                    dept_res = await self.db.execute(select(Department).where(func.lower(Department.name) == dept_key))
                    dept = dept_res.scalar_one_or_none()
                    if not dept:
                        code = "".join(w[0] for w in dept_name.split() if w[0].isalnum()).upper()
                        if not code:
                            code = dept_name[:3].upper()
                        
                        code_idx = 1
                        base_code = code
                        while True:
                            code_check = await self.db.execute(select(Department).where(Department.code == code))
                            if not code_check.scalar_one_or_none():
                                break
                            code = f"{base_code}{code_idx}"
                            code_idx += 1
                            
                        dept = Department(name=dept_name, code=code)
                        self.db.add(dept)
                        await self.db.flush()
                    depts_cache[dept_key] = dept.id
                dept_id = depts_cache[dept_key]

                # 2. Resolve role (case-insensitive)
                role_key = role_name.lower()
                if role_key not in roles_cache:
                    role_res = await self.db.execute(select(RoleManager).where(func.lower(RoleManager.name) == role_key))
                    role = role_res.scalar_one_or_none()
                    if not role:
                        val = role_name.lower().replace(" ", "_")
                        group_key = "verifier_general"
                        if "faculty" in val:
                            group_key = "faculty"
                        elif "head" in val or "hod" in val:
                            group_key = "hod"
                        elif "assistant" in val or "da" in val:
                            group_key = "verifier_da"
                        elif "superintendent" in val or "sp" in val:
                            group_key = "verifier_sp"
                        elif "director" in val or "registrar" in val:
                            group_key = "apex_approver"
                        elif "admin" in val:
                            group_key = "admin"
                            
                        role = RoleManager(name=role_name, value=val, group_key=group_key)
                        self.db.add(role)
                        await self.db.flush()
                    roles_cache[role_key] = role.id
                role_id = roles_cache[role_key]

                # 3. Create or update user
                user_res = await self.db.execute(select(User).where(User.email == email.lower()))
                existing_user = user_res.scalar_one_or_none()
                if existing_user:
                    existing_user.name = name
                    existing_user.department_id = dept_id
                    existing_user.role_id = role_id
                else:
                    new_user = User(
                        name=name,
                        email=email.lower(),
                        hashed_password=default_pw_hash,
                        department_id=dept_id,
                        role_id=role_id,
                        is_active=True,
                    )
                    self.db.add(new_user)
                
                await self.db.flush()
                success_count += 1
            except Exception as ex:
                errors.append(f"Row {row_num}: Database error {str(ex)}")

        await self.db.commit()
        return {
            "success": len(errors) == 0,
            "imported": success_count,
            "errors": errors
        }

    async def import_budget_csv(
        self, file_content: bytes, filename: str, financial_year_id: Optional[int] = None
    ) -> Dict[str, Any]:
        """
        Parses a CSV file containing budget details, resolves the financial year and departments,
        and saves budget allocations to the master records.
        """
        if not filename.endswith('.csv'):
            raise HTTPException(status_code=400, detail="Only CSV files (.csv) are supported")

        try:
            csv_text = file_content.decode("utf-8")
            reader = csv.reader(io.StringIO(csv_text))
            rows = list(reader)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse CSV file: {str(e)}")

        if not rows:
            raise HTTPException(status_code=400, detail="The CSV file is empty")

        # Normalize headers
        normalized_headers = [h.strip().lower().replace("_", " ") for h in rows[0]]

        def find_idx(keywords, required=False):
            for i, h in enumerate(normalized_headers):
                if any(kw in h for kw in keywords):
                    return i
            if required:
                raise HTTPException(status_code=400, detail=f"Required column matching one of {keywords} not found in CSV headers: {rows[0]}")
            return None

        dept_idx = find_idx(["department", "dept"], required=True)
        file_no_idx = find_idx(["file no", "file number", "file_no", "file"], required=True)
        item_idx = find_idx(["item name", "item_name", "procurement", "item", "details", "description"], required=True)
        
        unit_cost_idx = find_idx(["unit cost", "unit price", "rate", "cost"])
        qty_idx = find_idx(["quantity", "qty"])
        total_cost_idx = find_idx(["total cost", "budget amount", "total amount", "amount", "total"])
        
        exp_cat_idx = find_idx(["source of fund", "source_of_fund", "expenditure category", "expenditure type", "expenditure_category"])
        cat_idx = find_idx(["purchase category", "category", "type"])
        course_idx = find_idx(["course code", "course"])
        fy_idx = find_idx(["financial year", "fy", "financial_year"])

        # Extract unique department codes/names from CSV and validate they exist
        csv_depts = set()
        for row in rows[1:]:
            if not row or not any(row):
                continue
            if dept_idx < len(row):
                dept_code = str(row[dept_idx]).strip().lower()
                if dept_code:
                    csv_depts.add(dept_code)

        unrecognized_depts = []
        for dept_code in csv_depts:
            dept_res = await self.db.execute(
                select(Department).where(
                    (func.lower(Department.short_code) == dept_code) | 
                    (func.lower(Department.name) == dept_code)
                )
            )
            if not dept_res.scalar_one_or_none():
                unrecognized_depts.append(dept_code.upper())

        if unrecognized_depts:
            raise HTTPException(
                status_code=422,
                detail=f"Unrecognized departments in CSV: {', '.join(unrecognized_depts)}"
            )

        # Resolve financial year from form param if provided
        financial_year = None
        if financial_year_id is not None:
            fy_result = await self.db.execute(select(FinancialYear).where(FinancialYear.id == financial_year_id))
            financial_year = fy_result.scalar_one_or_none()
            if not financial_year:
                raise HTTPException(status_code=400, detail=f"Financial year with ID {financial_year_id} not found")

        # Fallback to active/default if not resolved yet
        if not financial_year:
            now = datetime.utcnow()
            fy_result = await self.db.execute(
                select(FinancialYear).where(
                    and_(FinancialYear.start_date <= now.date(), FinancialYear.end_date >= now.date())
                )
            )
            financial_year = fy_result.scalar_one_or_none()
            if not financial_year:
                fy_result = await self.db.execute(select(FinancialYear).where(FinancialYear.is_active == True).limit(1))
                financial_year = fy_result.scalar_one_or_none()
            if not financial_year:
                raise HTTPException(status_code=400, detail="No active financial year configured in the system")

        if financial_year.is_closed:
            raise HTTPException(
                status_code=400,
                detail=f"The financial year '{financial_year.label}' is closed. Budgets in closed financial years cannot be imported."
            )

        def clean_float(val_str) -> float:
            if not val_str:
                return 0.0
            cleaned = re.sub(r'[^\d.]', '', val_str)
            return float(cleaned) if cleaned else 0.0

        def clean_int(val_str) -> int:
            if not val_str:
                return 0
            cleaned = re.sub(r'[^\d]', '', val_str)
            return int(cleaned) if cleaned else 0

        success_count = 0
        errors = []
        depts_cache = {}

        for row_num, row in enumerate(rows[1:], start=2):
            if not row or not any(row):
                continue
            
            try:
                dept_code = str(row[dept_idx]).strip().upper()
                file_no = str(row[file_no_idx]).strip().upper()
                item_name = str(row[item_idx]).strip()

                if not dept_code or not file_no or not item_name:
                    errors.append(f"Row {row_num}: Missing required field values (Department, File No, Item Name)")
                    continue

                # Determine financial year for this row
                row_fy = financial_year
                if fy_idx is not None and fy_idx < len(row) and row[fy_idx]:
                    fy_label = str(row[fy_idx]).strip()
                    if fy_label:
                        fy_res = await self.db.execute(select(FinancialYear).where(FinancialYear.label == fy_label))
                        fy_obj = fy_res.scalar_one_or_none()
                        if fy_obj:
                            row_fy = fy_obj

                if row_fy.is_closed:
                    errors.append(f"Row {row_num}: Financial Year '{row_fy.label}' is closed. Budgets in closed financial years cannot be modified.")
                    continue

                # Parse amounts and quantities
                unit_cost = 0.0
                quantity = 1
                total_cost = 0.0

                has_unit_cost = unit_cost_idx is not None and unit_cost_idx < len(row) and row[unit_cost_idx]
                has_qty = qty_idx is not None and qty_idx < len(row) and row[qty_idx]
                has_total = total_cost_idx is not None and total_cost_idx < len(row) and row[total_cost_idx]

                if has_unit_cost:
                    unit_cost = clean_float(str(row[unit_cost_idx]))
                if has_qty:
                    quantity = clean_int(str(row[qty_idx]))
                    if quantity <= 0:
                        quantity = 1
                if has_total:
                    total_cost = clean_float(str(row[total_cost_idx]))

                # Calculate missing values
                if has_unit_cost and has_qty:
                    calculated_total = unit_cost * quantity
                    if not has_total or total_cost == 0.0:
                        total_cost = round(calculated_total, 2)
                    elif unit_cost == 0.0 and total_cost > 0.0:
                        unit_cost = total_cost / quantity
                elif has_total:
                    if not has_unit_cost or unit_cost == 0.0:
                        unit_cost = total_cost
                    if not has_qty:
                        quantity = 1

                exp_cat = "CAPEX"
                if exp_cat_idx is not None and exp_cat_idx < len(row) and row[exp_cat_idx]:
                    exp_cat = str(row[exp_cat_idx]).strip()

                cat = "equipment"
                if cat_idx is not None and cat_idx < len(row) and row[cat_idx]:
                    cat = str(row[cat_idx]).strip()

                course_code = "N/A"
                if course_idx is not None and course_idx < len(row) and row[course_idx]:
                    course_code = str(row[course_idx]).strip()

                dept_key = dept_code.lower()
                if dept_key not in depts_cache:
                    dept_res = await self.db.execute(
                        select(Department).where(
                            (func.lower(Department.short_code) == dept_key) | 
                            (func.lower(Department.name) == dept_key)
                        )
                    )
                    dept = dept_res.scalar_one_or_none()
                    if not dept:
                        raise HTTPException(status_code=422, detail=f"Department {dept_code} not found")
                    depts_cache[dept_key] = dept.id
                dept_id = depts_cache[dept_key]

                bm_res = await self.db.execute(
                    select(BudgetMaster).where(
                        and_(
                            BudgetMaster.department_id == dept_id,
                            BudgetMaster.file_no == file_no
                        )
                    )
                )
                bm = bm_res.scalar_one_or_none()

                if bm:
                    bm_fy_res = await self.db.execute(select(FinancialYear).where(FinancialYear.id == bm.financial_year_id))
                    bm_fy = bm_fy_res.scalar_one_or_none()
                    if bm_fy and bm_fy.is_closed:
                        errors.append(f"Row {row_num}: Budget file {file_no} already exists in a closed Financial Year '{bm_fy.label}' and cannot be modified.")
                        continue

                    bm.total_cost = total_cost
                    bm.unit_cost = unit_cost
                    bm.item_name = item_name
                    bm.quantity = quantity
                    bm.financial_year_id = row_fy.id
                    if exp_cat_idx is not None:
                        bm.source_of_fund = exp_cat
                    if cat_idx is not None:
                        bm.category = cat
                    if course_idx is not None:
                        bm.course_code = course_code
                else:
                    bm = BudgetMaster(
                        department_id=dept_id,
                        financial_year_id=row_fy.id,
                        source_of_fund=exp_cat,
                        item_name=item_name,
                        category=cat,
                        course_code=course_code,
                        unit_cost=round(unit_cost, 2),
                        quantity=quantity,
                        total_cost=round(total_cost, 2),
                        file_no=file_no,
                        is_revision=False,
                    )
                    self.db.add(bm)

                success_count += 1
            except Exception as ex:
                errors.append(f"Row {row_num}: Error: {str(ex)}")

        await self.db.commit()
        return {
            "success": len(errors) == 0,
            "imported": success_count,
            "errors": errors
        }
