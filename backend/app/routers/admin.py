from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query, Request
from sqlalchemy.ext.asyncio import AsyncSession
from app.core.limiter import limiter
from sqlalchemy import select, func, and_, or_, text
from sqlalchemy.orm import selectinload
from datetime import datetime
from typing import Optional, List, Any
from app.core.config import settings

from app.core.database import get_db
from app.core.deps import require_roles, get_current_user
from app.core.security import get_password_hash
from app.models.user import User, Department, RoleManager
from app.models.budget import BudgetMaster, FinancialYear, PurchaseCategory, ProcurementManager, PhaseManager, Settings, SourceOfFund
from app.models.purchase_request import WorkFlowHierarchy, PurchaseRequest, PurchaseRequestItem, RequestStatus, PurchaseRequestHistory, PurchaseRequestFlow

router = APIRouter(prefix="/api/admin", tags=["admin"])
AdminDep = Depends(require_roles("admin"))
DeanOrAdminDep = Depends(require_roles("admin", "dean_approver", "apex_approver"))
async def check_dean_budget_write(user=Depends(get_current_user)):
    if user.role.group_key not in ("dean_approver", "admin"):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    return user

DeanBudgetDep = Depends(check_dean_budget_write)
BudgetViewDep = Depends(require_roles("admin", "dean_approver", "hod", "faculty", "apex_approver"))


# ─────────────────────────────────────────────────────────────────────────────
# USERS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/users")
async def list_users(db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(User).order_by(User.name))
    users = result.scalars().all()
    return [
        {
            "id": u.id,
            "title": u.title,
            "name": u.name,
            "email": u.email,
            "designation": u.designation,
            "gender": u.gender,
            "role_id": u.role_id,
            "department_id": u.department_id,
            "is_active": u.is_active,
            "is_approved": u.is_approved,
            "last_login_at": u.last_login_at.isoformat() if u.last_login_at else None,
            "created_at": u.created_at.isoformat() if u.created_at else None,
        }
        for u in users
    ]


@router.post("/users")
async def create_user(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    existing = await db.execute(select(User).where(User.email == body["email"].lower()))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Email already in use")
    u = User(
        title=body.get("title", "Mr."),
        name=body["name"],
        email=body["email"].lower(),
        hashed_password=get_password_hash(body["password"]),
        designation=body.get("designation", ""),
        gender=body.get("gender", "male"),
        role_id=body.get("role_id"),
        department_id=body.get("department_id"),
        is_active=True,
        is_approved=body.get("is_approved", True),
    )
    db.add(u)
    await db.commit()
    return {"message": "User created", "id": u.id}


@router.get("/users/import-template")
async def import_template(_=AdminDep):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Users Import"

    # Set headers
    headers = ["Name", "Email", "Department Name", "Role Name"]
    ws.append(headers)

    # Styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Dark Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers))):
        for cell in col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

    # Sample rows
    samples = [
        ["John Doe", "john.doe@nitt.edu", "Computer Science and Engineering", "Faculty"],
        ["Jane Smith", "jane.smith@nitt.edu", "Electronics and Communication Engineering", "Head of Department"],
    ]
    for row in samples:
        ws.append(row)

    # Style samples
    for row_idx in range(2, len(samples) + 2):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Calibri", size=11)
            cell.border = thin_border

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 15)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="users_import_template.xlsx"'}
    )


@router.post("/users/import")
async def import_users(file: UploadFile, db: AsyncSession = Depends(get_db), _=AdminDep):
    from app.services.import_service import ImportService
    content = await file.read()
    import_service = ImportService(db)
    return await import_service.import_users(content, file.filename)


@router.put("/users/{user_id}")
async def update_user(user_id: int, body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(User).where(User.id == user_id))
    u = result.scalar_one_or_none()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    if "name" in body:
        u.name = body["name"]
    if "email" in body:
        new_email = body["email"].lower()
        if new_email != u.email:
            existing = await db.execute(select(User).where(User.email == new_email))
            if existing.scalar_one_or_none():
                raise HTTPException(status_code=409, detail="Email already in use")
            u.email = new_email
    if "password" in body and body["password"]:
        u.hashed_password = get_password_hash(body["password"])
    if "title" in body:
        u.title = body["title"]
    if "designation" in body:
        u.designation = body["designation"]
    if "role_id" in body:
        u.role_id = body["role_id"]
    if "department_id" in body:
        u.department_id = body["department_id"]
    if "is_active" in body:
        u.is_active = bool(body["is_active"])
    if "is_approved" in body:
        u.is_approved = bool(body["is_approved"])
    await db.commit()
    return {"message": "User updated"}


@router.post("/users/{user_id}/reset-password")
async def reset_password(user_id: int, body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(User).where(User.id == user_id))
    u = result.scalar_one_or_none()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    new_password = body.get("password", "Password@123")
    u.hashed_password = get_password_hash(new_password)
    await db.commit()
    return {"message": "Password reset successfully"}


# ─────────────────────────────────────────────────────────────────────────────
# DEPARTMENTS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/departments")
async def list_departments(db: AsyncSession = Depends(get_db), _=BudgetViewDep):
    result = await db.execute(select(Department).order_by(Department.short_code))
    return [{"id": d.id, "name": d.name, "short_code": d.short_code} for d in result.scalars()]


@router.post("/departments")
async def create_department(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    d = Department(name=body["name"], short_code=body["short_code"].upper())
    db.add(d)
    await db.commit()
    return {"message": "Department created", "id": d.id}


# ─────────────────────────────────────────────────────────────────────────────
# ROLES
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/roles")
async def list_roles(db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(RoleManager).order_by(RoleManager.name))
    return [{"id": r.id, "name": r.name, "value": r.value, "group_key": r.group_key} for r in result.scalars()]


@router.post("/roles")
async def create_role(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    name = body.get("name")
    value = body.get("value")
    group_key = body.get("group_key")
    if not name or not value or not group_key:
        raise HTTPException(status_code=400, detail="Missing name, value, or group_key")
    
    res = await db.execute(select(RoleManager).where(RoleManager.value == value))
    existing = res.scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=400, detail=f"Role value '{value}' already exists")
        
    role = RoleManager(name=name, value=value, group_key=group_key)
    db.add(role)
    await db.commit()
    return {"message": "Role created", "id": role.id}


# ─────────────────────────────────────────────────────────────────────────────
# FINANCIAL YEARS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/financial-years")
async def list_financial_years(db: AsyncSession = Depends(get_db), _=BudgetViewDep):
    result = await db.execute(select(FinancialYear).order_by(FinancialYear.start_date.desc()))
    return [{"id": fy.id, "label": fy.label, "start_date": fy.start_date.isoformat(), "end_date": fy.end_date.isoformat(), "is_active": fy.is_active} for fy in result.scalars()]


@router.post("/financial-years")
async def create_financial_year(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    from datetime import date as dateobj
    fy = FinancialYear(
        label=body["label"],
        start_date=dateobj.fromisoformat(body["start_date"]),
        end_date=dateobj.fromisoformat(body["end_date"]),
        is_active=body.get("is_active", True),
        is_closed=body.get("is_closed", False),
    )
    db.add(fy)
    await db.commit()
    return {"message": "Financial year created", "id": fy.id}


@router.get("/financial-years/rollover-candidates")
async def get_rollover_candidates(db: AsyncSession = Depends(get_db), _=DeanBudgetDep):
    """
    Returns all unused budget files from the currently active financial year.
    Unused = committed_amount == 0 AND utilized_amount == 0 (no PRs have touched them).
    These are the files the Dean can choose to carry over on rollover.
    """
    active_fy_res = await db.execute(select(FinancialYear).where(FinancialYear.is_active == True))
    active_fy = active_fy_res.scalar_one_or_none()
    if not active_fy:
        return []

    from app.models.user import Department
    unused_res = await db.execute(
        select(BudgetMaster)
        .options(selectinload(BudgetMaster.department))
        .where(
            BudgetMaster.financial_year_id == active_fy.id,
            BudgetMaster.committed_amount == 0.0,
            BudgetMaster.utilized_amount == 0.0,
        )
        .order_by(BudgetMaster.department_id, BudgetMaster.file_no)
    )
    candidates = unused_res.scalars().all()
    return [
        {
            "id": b.id,
            "file_no": b.file_no,
            "item_name": b.item_name,
            "category": b.category,
            "source_of_fund": b.source_of_fund,
            "total_allocation": b.total_allocation,
            "unit_cost": b.unit_cost,
            "quantity": b.quantity,
            "department_id": b.department_id,
            "department_code": b.department.short_code if b.department else str(b.department_id),
            "financial_year_id": b.financial_year_id,
            "fy_label": active_fy.label,
            "is_temporary": b.file_no.upper().startswith("TEMP/"),
        }
        for b in candidates
    ]


@router.post("/financial-years/rollover")
async def financial_year_rollover(
    body: dict = {},
    db: AsyncSession = Depends(get_db),
    _=DeanBudgetDep
):
    """
    Rolls the active financial year to the next one.
    Optional body field `budget_file_ids` (list[int]): if provided, only those
    unused budget files are carried over to the new FY with revised (R-prefix)
    file numbers. If omitted or empty, ALL unused files are carried over.
    """
    selected_budget_file_ids: list = body.get("budget_file_ids", [])
    # None/empty means roll ALL; otherwise restrict to provided IDs
    # 1. Find the current active financial year
    active_fy_stmt = select(FinancialYear).where(FinancialYear.is_active == True)
    active_fy_res = await db.execute(active_fy_stmt)
    active_fy = active_fy_res.scalar_one_or_none()
    if not active_fy:
        raise HTTPException(status_code=400, detail="No active financial year found to rollover.")

    # 2. Parse current active FY label to calculate the next FY label
    import re
    match = re.match(r"^(\d{4})-(\d{2})$", active_fy.label)
    if not match:
        raise HTTPException(status_code=400, detail=f"Invalid financial year label format: {active_fy.label}")
    
    start_year = int(match.group(1))
    end_year_short = int(match.group(2))
    
    next_start_year = start_year + 1
    next_end_year_short = (end_year_short + 1) % 100
    next_label = f"{next_start_year}-{next_end_year_short:02d}"
    
    from datetime import date
    next_start_date = date(next_start_year, 4, 1)
    next_end_date = date(next_start_year + 1, 3, 31)

    # Find if next FY already exists
    next_fy_stmt = select(FinancialYear).where(FinancialYear.label == next_label)
    next_fy_res = await db.execute(next_fy_stmt)
    next_fy = next_fy_res.scalar_one_or_none()
    
    if next_fy:
        if next_fy.is_closed:
            raise HTTPException(status_code=400, detail=f"The next financial year '{next_label}' is already closed.")
        next_fy.is_active = True
    else:
        next_fy = FinancialYear(
            label=next_label,
            start_date=next_start_date,
            end_date=next_end_date,
            is_active=True,
            is_closed=False
        )
        db.add(next_fy)
        await db.flush()

    # Close current active FY
    active_fy.is_active = False
    active_fy.is_closed = True

    # 3. Find all active/in-progress PRs in the closed FY
    from app.models.purchase_request import PurchaseRequest, RequestStatus, PurchaseRequestItem, PurchaseRequestFlow, PurchaseRequestAssignment, TechnicalEvaluation, FinancialEvaluation, CommercialEvaluation, Document, PRReferral, PurchaseRequestHistory
    
    active_prs_res = await db.execute(
        select(PurchaseRequest)
        .options(
            selectinload(PurchaseRequest.items).selectinload(PurchaseRequestItem.budget_file),
            selectinload(PurchaseRequest.flow),
            selectinload(PurchaseRequest.assignments),
            selectinload(PurchaseRequest.technical_evaluations),
            selectinload(PurchaseRequest.financial_evaluations),
            selectinload(PurchaseRequest.commercial_evaluations),
            selectinload(PurchaseRequest.documents),
            selectinload(PurchaseRequest.referrals),
            selectinload(PurchaseRequest.initiator).selectinload(User.department)
        )
        .where(
            PurchaseRequest.financial_year_id == active_fy.id,
            PurchaseRequest.current_status.in_([
                RequestStatus.PR_SUBMITTED,
                RequestStatus.IN_PROGRESS,
                RequestStatus.SENT_BACK
            ])
        )
    )
    old_prs = active_prs_res.scalars().all()

    from app.services.budget_service import BudgetService
    budget_svc = BudgetService(db)
    dept_seqs = {}

    for old_pr in old_prs:
        # Unlock budget reservations in old FY
        await budget_svc.unlock_amount(old_pr)

        # For each item, resolve or clone its budget file to the new FY
        for item in old_pr.items:
            old_bm = item.budget_file
            if old_bm:
                new_bm_file_no = old_bm.file_no.replace(active_fy.label, next_fy.label)
                new_bm_file_no = re.sub(r'F\.No\.', 'R', new_bm_file_no, flags=re.IGNORECASE)
                new_bm_file_no_alt = old_bm.file_no.replace(active_fy.label.lower(), next_fy.label.lower())
                new_bm_file_no_alt = re.sub(r'F\.No\.', 'R', new_bm_file_no_alt, flags=re.IGNORECASE)
                
                bm_find = await db.execute(
                    select(BudgetMaster).where(
                        BudgetMaster.financial_year_id == next_fy.id,
                        BudgetMaster.file_no.in_([new_bm_file_no, new_bm_file_no_alt, old_bm.file_no])
                    )
                )
                new_bm = bm_find.scalar_one_or_none()
                if not new_bm:
                    new_bm = BudgetMaster(
                        department_id=old_bm.department_id,
                        financial_year_id=next_fy.id,
                        source_of_fund=old_bm.source_of_fund,
                        item_name=old_bm.item_name,
                        category=old_bm.category,
                        course_code=old_bm.course_code,
                        unit_cost=old_bm.unit_cost,
                        quantity=old_bm.quantity,
                        total_allocation=old_bm.total_allocation,
                        file_no=new_bm_file_no,
                        is_revision=old_bm.is_revision,
                        expert1_id=old_bm.expert1_id,
                        expert2_id=old_bm.expert2_id,
                        director_faculty_id=old_bm.director_faculty_id,
                        committed_amount=0.0,
                        utilized_amount=0.0
                    )
                    db.add(new_bm)
                    await db.flush()

        # Generate revised reference sequence number per department
        dept_id = old_pr.initiator.department_id
        if dept_id not in dept_seqs:
            q = await db.execute(
                select(func.count(PurchaseRequest.id))
                .join(User, PurchaseRequest.initiator_id == User.id)
                .where(
                    PurchaseRequest.financial_year_id == next_fy.id,
                    PurchaseRequest.parent_pr_id.isnot(None),
                    User.department_id == dept_id
                )
            )
            dept_seqs[dept_id] = q.scalar() or 0
        
        dept_seqs[dept_id] += 1
        seq_num = dept_seqs[dept_id]
        dept_code = old_pr.initiator.department.short_code if old_pr.initiator.department else "GEN"
        revised_ref = f"PR/{dept_code}/{next_fy.label}/R-{seq_num:02d}"

        # Clone the Purchase Request
        new_pr = PurchaseRequest(
            category_id=old_pr.category_id,
            financial_year_id=next_fy.id,
            initiator_id=old_pr.initiator_id,
            nominee_id=old_pr.nominee_id,
            procurement_id=old_pr.procurement_id,
            purchase_type=old_pr.purchase_type,
            amount=old_pr.amount,
            emd=old_pr.emd,
            performance_security=old_pr.performance_security,
            current_status=old_pr.current_status,
            vendor_list_link=old_pr.vendor_list_link,
            is_item_split=old_pr.is_item_split,
            item_split_justification=old_pr.item_split_justification,
            is_quantity_split=old_pr.is_quantity_split,
            quantity_split_details=old_pr.quantity_split_details,
            is_service_center_in_south=old_pr.is_service_center_in_south,
            service_center_south_desc=old_pr.service_center_south_desc,
            basis_of_estimate_details=old_pr.basis_of_estimate_details,
            delivery_mode=old_pr.delivery_mode,
            delivery_location=old_pr.delivery_location,
            exemption=old_pr.exemption,
            exemption_remarks=old_pr.exemption_remarks,
            is_training_required=old_pr.is_training_required,
            training_type=old_pr.training_type,
            training_vendor=old_pr.training_vendor,
            training_comments=old_pr.training_comments,
            tender_reference_number=old_pr.tender_reference_number,
            date_of_tender=old_pr.date_of_tender,
            date_of_tech_bid_opening=old_pr.date_of_tech_bid_opening,
            date_of_financial_bid_opening=old_pr.date_of_financial_bid_opening,
            aa_approved_at=old_pr.aa_approved_at,
            te_initiated_at=old_pr.te_initiated_at,
            te_approved_at=old_pr.te_approved_at,
            fs_initiated_at=old_pr.fs_initiated_at,
            fs_approved_at=old_pr.fs_approved_at,
            po_initiated_at=old_pr.po_initiated_at,
            po_approved_at=old_pr.po_approved_at,
            faculty1_id=old_pr.faculty1_id,
            faculty2_id=old_pr.faculty2_id,
            faculty3_id=old_pr.faculty3_id,
            aa_approver_id=old_pr.aa_approver_id,
            form_data=old_pr.form_data,
            parent_pr_id=old_pr.id,
            lpc_remarks=old_pr.lpc_remarks,
            lpc_committee_members=old_pr.lpc_committee_members,
            lpc_minutes_reference=old_pr.lpc_minutes_reference,
            single_bid_justification=old_pr.single_bid_justification,
            icr_number=revised_ref
        )
        db.add(new_pr)
        await db.flush()

        # Clone items, linking to new budgets
        for item in old_pr.items:
            old_item_bm = item.budget_file
            new_item_bm = None
            if old_item_bm:
                new_item_bm_file_no = old_item_bm.file_no.replace(active_fy.label, next_fy.label)
                new_item_bm_file_no = re.sub(r'F\.No\.', 'R', new_item_bm_file_no, flags=re.IGNORECASE)
                new_item_bm_file_no_alt = old_item_bm.file_no.replace(active_fy.label.lower(), next_fy.label.lower())
                new_item_bm_file_no_alt = re.sub(r'F\.No\.', 'R', new_item_bm_file_no_alt, flags=re.IGNORECASE)
                new_bm_find = await db.execute(
                    select(BudgetMaster).where(
                        BudgetMaster.financial_year_id == next_fy.id,
                        BudgetMaster.file_no.in_([new_item_bm_file_no, new_item_bm_file_no_alt, old_item_bm.file_no])
                    )
                )
                new_item_bm = new_bm_find.scalar_one_or_none()

            new_item = PurchaseRequestItem(
                purchase_request_id=new_pr.id,
                budget_file_id=new_item_bm.id if new_item_bm else None,
                item_description=item.item_description,
                quantity=item.quantity,
                estimated_total=item.estimated_total,
                charges=item.charges,
                requirement_type=item.requirement_type,
                availability=item.availability,
                availability_remarks=item.availability_remarks,
                site_readiness=item.site_readiness,
                site_readiness_remarks=item.site_readiness_remarks,
                warranty=item.warranty,
                delivery_period=item.delivery_period,
                present_stock=item.present_stock,
                justification_for_procurement=item.justification_for_procurement,
                previous_file_no_reference=item.previous_file_no_reference,
                installation_required=item.installation_required,
                tech_specs_text=item.tech_specs_text,
                gem_link=item.gem_link,
            )
            db.add(new_item)
        await db.flush()

        # Clone flow state
        if old_pr.flow:
            new_flow = PurchaseRequestFlow(
                purchase_request_id=new_pr.id,
                phase_id=old_pr.flow.phase_id,
                step_order=old_pr.flow.step_order,
                rejected=old_pr.flow.rejected
            )
            db.add(new_flow)

        # Clone assignments
        for ass in old_pr.assignments:
            new_ass = PurchaseRequestAssignment(
                purchase_request_id=new_pr.id,
                assigned_by_id=ass.assigned_by_id,
                assigned_da_id=ass.assigned_da_id,
                status=ass.status,
                assigned_at=ass.assigned_at
            )
            db.add(new_ass)

        # Clone evaluations
        for te in old_pr.technical_evaluations:
            new_te = TechnicalEvaluation(
                purchase_request_id=new_pr.id,
                vendor_name=te.vendor_name,
                is_qualified=te.is_qualified,
                remarks=te.remarks,
                created_at=te.created_at
            )
            db.add(new_te)
        
        for fe in old_pr.financial_evaluations:
            new_fe = FinancialEvaluation(
                purchase_request_id=new_pr.id,
                vendor_name=fe.vendor_name,
                quoted_amount=fe.quoted_amount,
                ranking=fe.ranking,
                is_awarded=fe.is_awarded,
                remarks=fe.remarks,
                unit_price=fe.unit_price,
                taxes=fe.taxes,
                delivery_period=fe.delivery_period,
                warranty=fe.warranty,
                created_at=fe.created_at
            )
            db.add(new_fe)
            
        for ce in old_pr.commercial_evaluations:
            new_ce = CommercialEvaluation(
                purchase_request_id=new_pr.id,
                vendor_name=ce.vendor_name,
                vendor_email=ce.vendor_email,
                quoted_amount=ce.quoted_amount,
                is_qualified=ce.is_qualified,
                remarks=ce.remarks
            )
            db.add(new_ce)

        # Clone documents
        for doc in old_pr.documents:
            new_doc = Document(
                purchase_request_id=new_pr.id,
                doc_key=doc.doc_key,
                doc_value=doc.doc_value,
                uploaded_by_id=doc.uploaded_by_id,
                updated_at=doc.updated_at
            )
            db.add(new_doc)

        # Clone referrals
        for ref in old_pr.referrals:
            new_ref = PRReferral(
                purchase_request_id=new_pr.id,
                referred_by_id=ref.referred_by_id,
                referred_to_id=ref.referred_to_id,
                query=ref.query,
                query_document_path=ref.query_document_path,
                response=ref.response,
                response_document_path=ref.response_document_path,
                status=ref.status,
                created_at=ref.created_at,
                responded_at=ref.responded_at
            )
            db.add(new_ref)

        # Lock budget amounts on the new cloned budget files
        await budget_svc.lock_amount(new_pr)

        # Mark old PR as rolled_over
        old_pr.current_status = RequestStatus.ROLLED_OVER

        # Add history entries
        old_history = PurchaseRequestHistory(
            purchase_request_id=old_pr.id,
            status="Rolled Over",
            remarks=f"Purchase request rolled over to next financial year {next_fy.label}. Revised reference: {revised_ref}",
            acted_at=datetime.utcnow()
        )
        db.add(old_history)

        new_history = PurchaseRequestHistory(
            purchase_request_id=new_pr.id,
            status="Rolled Over",
            remarks=f"Purchase request rolled over from financial year {active_fy.label}. Original reference: {old_pr.icr_number}",
            acted_at=datetime.utcnow()
        )
        db.add(new_history)

    # 4. Find unused budget files in the closed FY (committed=0, utilized=0).
    # If the request body supplied specific budget_file_ids, only roll those.
    # Otherwise, roll ALL unused files (legacy / default behaviour).
    unused_filter = [
        BudgetMaster.financial_year_id == active_fy.id,
        BudgetMaster.committed_amount == 0.0,
        BudgetMaster.utilized_amount == 0.0,
    ]
    if selected_budget_file_ids:
        unused_filter.append(BudgetMaster.id.in_(selected_budget_file_ids))

    unused_bms_res = await db.execute(select(BudgetMaster).where(*unused_filter))
    unused_bms = unused_bms_res.scalars().all()
    rolled_file_count = 0
    for old_bm in unused_bms:
        # Derive the revised file number:
        # Replace the FY label and rename F.No. → R to mark as revised
        new_bm_file_no = old_bm.file_no.replace(active_fy.label, next_fy.label)
        new_bm_file_no = re.sub(r'F\.No\.', 'R', new_bm_file_no, flags=re.IGNORECASE)

        new_bm_file_no_alt = old_bm.file_no.replace(active_fy.label.lower(), next_fy.label.lower())
        new_bm_file_no_alt = re.sub(r'F\.No\.', 'R', new_bm_file_no_alt, flags=re.IGNORECASE)

        # Check if already cloned / exists in new FY
        bm_find = await db.execute(
            select(BudgetMaster).where(
                BudgetMaster.financial_year_id == next_fy.id,
                BudgetMaster.file_no.in_([new_bm_file_no, new_bm_file_no_alt, old_bm.file_no])
            )
        )
        new_bm = bm_find.scalar_one_or_none()
        if not new_bm:
            new_bm = BudgetMaster(
                department_id=old_bm.department_id,
                financial_year_id=next_fy.id,
                source_of_fund=old_bm.source_of_fund,
                item_name=old_bm.item_name,
                category=old_bm.category,
                course_code=old_bm.course_code,
                unit_cost=old_bm.unit_cost,
                quantity=old_bm.quantity,
                total_allocation=old_bm.total_allocation,
                file_no=new_bm_file_no,
                is_revision=True,  # Mark as a revised/rolled-over file
                expert1_id=old_bm.expert1_id,
                expert2_id=old_bm.expert2_id,
                director_faculty_id=old_bm.director_faculty_id,
                committed_amount=0.0,
                utilized_amount=0.0,
                remarks=old_bm.remarks
            )
            db.add(new_bm)
            rolled_file_count += 1
    await db.flush()

    await db.commit()
    return {
        "message": "Financial Year rollover completed successfully.",
        "closed_year": active_fy.label,
        "opened_year": next_fy.label,
        "rolled_over_pr_count": len(old_prs),
        "rolled_over_file_count": rolled_file_count,
    }


# ─────────────────────────────────────────────────────────────────────────────
# SETTINGS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/settings")
async def get_settings(db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(Settings))
    return {s.key_name: s.value for s in result.scalars()}


@router.put("/settings/{key}")
async def update_setting(key: str, body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(Settings).where(Settings.key_name == key))
    setting = result.scalar_one_or_none()
    if setting:
        setting.value = body["value"]
    else:
        setting = Settings(key_name=key, value=body["value"])
        db.add(setting)
    await db.commit()
    return {"message": "Setting updated"}


# ─────────────────────────────────────────────────────────────────────────────
# BUDGET
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/budget")
async def list_budget(
    skip: int = 0,
    limit: int = Query(default=50, le=2000),
    search: Optional[str] = None,
    department_id: Optional[int] = None,
    financial_year_id: Optional[int] = None,
    is_temporary: Optional[bool] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=BudgetViewDep
):
    base_query = select(BudgetMaster)
    
    group_key = user.role.group_key if user.role else None
    filters = []
    
    is_temp = is_temporary if isinstance(is_temporary, bool) else None
    if is_temp is not None:
        if is_temp:
            filters.append(BudgetMaster.file_no.ilike("TEMP/%"))
        else:
            filters.append(~BudgetMaster.file_no.ilike("TEMP/%"))
            
    if financial_year_id is not None:
        filters.append(BudgetMaster.financial_year_id == financial_year_id)
        
    if search:
        search_pattern = f"%{search}%"
        filters.append(
            or_(
                BudgetMaster.file_no.ilike(search_pattern),
                BudgetMaster.item_name.ilike(search_pattern)
            )
        )

    # Scopes
    if group_key in ("hod", "faculty") and user.department_id:
        filters.append(BudgetMaster.department_id == user.department_id)
        
    if group_key == "faculty":
        from sqlalchemy import exists
        from app.models.purchase_request import PurchaseRequestItem, PurchaseRequest, PurchaseRequestHistory
        
        pr_budget_exists = exists().where(
            and_(
                PurchaseRequestItem.budget_file_id == BudgetMaster.id,
                PurchaseRequestItem.purchase_request_id == PurchaseRequest.id,
                or_(
                    PurchaseRequest.initiator_id == user.id,
                    exists().where(
                        and_(
                            PurchaseRequestHistory.purchase_request_id == PurchaseRequest.id,
                            PurchaseRequestHistory.current_approver_id == user.id
                        )
                    )
                )
            )
        )
        filters.append(
            or_(
                BudgetMaster.allocated_initiator_id == user.id,
                pr_budget_exists
            )
        )
    else:
        if department_id is not None:
            filters.append(BudgetMaster.department_id == department_id)

    if filters:
        base_query = base_query.where(and_(*filters))

    # Get total count matching criteria
    total = await db.scalar(select(func.count()).select_from(base_query.subquery())) or 0

    result = await db.execute(
        base_query
        .outerjoin(Department)
        .options(
            selectinload(BudgetMaster.expert1),
            selectinload(BudgetMaster.expert2),
            selectinload(BudgetMaster.director_faculty),
            selectinload(BudgetMaster.allocated_initiator)
        )
        .order_by(BudgetMaster.created_at.desc(), BudgetMaster.id.desc())
        .offset(skip)
        .limit(limit)
    )
    entries = result.scalars().all()
    items = [
        {
            "id": b.id, "item_name": b.item_name,
            "total_cost": b.total_allocation,
            "total_allocation": b.total_allocation,
            "locked_amount": b.committed_amount if b.committed_amount is not None else 0.0,
            "committed_amount": b.committed_amount if b.committed_amount is not None else 0.0,
            "deducted_amount": b.utilized_amount if b.utilized_amount is not None else 0.0,
            "utilized_amount": b.utilized_amount if b.utilized_amount is not None else 0.0,
            "available_amount": b.total_allocation - (b.committed_amount or 0.0) - (b.utilized_amount or 0.0),
            "available_balance": b.total_allocation - (b.committed_amount or 0.0) - (b.utilized_amount or 0.0),
            "department_id": b.department_id,
            "financial_year_id": b.financial_year_id, "source_of_fund": b.source_of_fund, "expenditure_category": b.source_of_fund,
            "category": b.category, "unit_cost": b.unit_cost, "quantity": b.quantity,
            "file_no": b.file_no,
            "remarks": b.remarks,
            "expert1_id": b.expert1_id,
            "expert2_id": b.expert2_id,
            "director_faculty_id": b.director_faculty_id,
            "allocated_initiator_id": b.allocated_initiator_id,
            "expert1": {"id": b.expert1.id, "name": b.expert1.name, "email": b.expert1.email} if b.expert1 else None,
            "expert2": {"id": b.expert2.id, "name": b.expert2.name, "email": b.expert2.email} if b.expert2 else None,
            "director_faculty": {"id": b.director_faculty.id, "name": b.director_faculty.name, "email": b.director_faculty.email} if b.director_faculty else None,
            "allocated_initiator": {"id": b.allocated_initiator.id, "name": b.allocated_initiator.name, "email": b.allocated_initiator.email} if b.allocated_initiator else None,
            "attachment_path": b.attachment_path,
            "attachment_url": f"/static/uploads/{b.attachment_path}" if b.attachment_path else None,
            "project_code": b.project_code,
            "principal_investigator": b.principal_investigator,
            "project_due_date": b.project_due_date.isoformat() if b.project_due_date else None,
            "created_at": b.created_at.isoformat() if b.created_at else None,
        }
        for b in entries
    ]
    return {"items": items, "total": total}


@router.get("/budget/summary")
async def budget_summary(db: AsyncSession = Depends(get_db), _=DeanOrAdminDep):
    """System-wide consolidated budget totals for admin dashboard."""
    result = await db.execute(select(BudgetMaster))
    entries = result.scalars().all()
    total = sum(b.total_cost for b in entries)
    locked = sum(b.locked_amount for b in entries)
    deducted = sum(b.deducted_amount for b in entries)
    return {"total": total, "locked": locked, "deducted": deducted, "available": total - locked - deducted}


async def get_stored_categories(db: AsyncSession):
    # Try new SourceOfFund master table first
    sof_result = await db.execute(select(SourceOfFund).where(SourceOfFund.is_active == True).order_by(SourceOfFund.name))
    sof_rows = sof_result.scalars().all()
    
    if sof_rows:
        exp_list = [s.name for s in sof_rows]
    else:
        # Fallback: read from legacy Settings table and seed SourceOfFund
        result_exp = await db.execute(select(Settings).where(Settings.key_name == "budget_source_of_fund_categories"))
        exp_setting = result_exp.scalar_one_or_none()
        if exp_setting:
            exp_list = [c.strip() for c in exp_setting.value.split(",") if c.strip()]
        else:
            exp_list = ["CAPEX (OH-35)", "REVEX (OH-31)", "HOSTEL", "NIMCET", "ID", "PMRF", "SEED-GRANT", "HEFA", "STUDENT-WELFARE", "R&C"]
            db.add(Settings(key_name="budget_source_of_fund_categories", value="CAPEX (OH-35),REVEX (OH-31),HOSTEL,NIMCET,ID,PMRF,SEED-GRANT,HEFA,STUDENT-WELFARE,R&C"))
            await db.flush()
        # Seed SourceOfFund table from the list
        for name in exp_list:
            db.add(SourceOfFund(name=name, is_active=True))
        await db.flush()

    result_item = await db.execute(select(Settings).where(Settings.key_name == "budget_item_categories"))
    item_setting = result_item.scalar_one_or_none()
    if item_setting:
        item_list = [c.strip() for c in item_setting.value.split(",") if c.strip()]
    else:
        item_list = ["computer", "lab_equipment", "software", "furniture"]
        db.add(Settings(key_name="budget_item_categories", value="computer,lab_equipment,software,furniture"))
        await db.flush()

    # Also load added_by_dean
    result_dean = await db.execute(select(Settings).where(Settings.key_name == "budget_categories_added_by_dean"))
    dean_setting = result_dean.scalar_one_or_none()
    if dean_setting:
        import json
        try:
            dean_cats = json.loads(dean_setting.value)
        except Exception:
            dean_cats = {"source_of_fund": [], "item": []}
    else:
        dean_cats = {"source_of_fund": [], "item": []}

    return {
        "source_of_fund_categories": exp_list,
        "expenditure_categories": exp_list,
        "item_categories": item_list,
        "added_by_dean": dean_cats
    }


@router.get("/budget/categories")
async def get_budget_categories(db: AsyncSession = Depends(get_db), _=BudgetViewDep):
    return await get_stored_categories(db)


@router.post("/budget/categories")
async def add_budget_category(
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    await db.refresh(current_user, ["role"])
    if current_user.role.group_key != "admin":
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    cat_type = body.get("type")
    val = body.get("value")
    if not val or not val.strip():
        raise HTTPException(status_code=400, detail="Category value is required")
    val = val.strip()

    if cat_type in ("source_of_fund", "expenditure"):
        key = "budget_source_of_fund_categories"
    elif cat_type == "item":
        key = "budget_item_categories"
    else:
        raise HTTPException(status_code=400, detail="Invalid type: must be 'source_of_fund' or 'item'")

    result = await db.execute(select(Settings).where(Settings.key_name == key))
    setting = result.scalar_one_or_none()
    
    already_exists = False
    
    if setting:
        existing = [c.strip() for c in setting.value.split(",") if c.strip()]
        if val in existing:
            already_exists = True
        else:
            existing.append(val)
            setting.value = ",".join(existing)
    else:
        if cat_type == "expenditure":
            defaults = ["CAPEX", "OPEX", val]
        else:
            defaults = ["computer", "lab_equipment", "software", "furniture", val]
        setting = Settings(key_name=key, value=",".join(defaults))
        db.add(setting)

    # Record if added by admin or dean budget role
    if not already_exists and current_user.role.group_key in ["admin", "dean_approver"]:
        result_dean = await db.execute(select(Settings).where(Settings.key_name == "budget_categories_added_by_dean"))
        dean_setting = result_dean.scalar_one_or_none()
        import json
        if dean_setting:
            try:
                dean_cats = json.loads(dean_setting.value)
            except Exception:
                dean_cats = {"expenditure": [], "item": []}
            
            if cat_type in ("source_of_fund", "expenditure"):
                if "source_of_fund" not in dean_cats:
                    dean_cats["source_of_fund"] = []
                if val not in dean_cats["source_of_fund"]:
                    dean_cats["source_of_fund"].append(val)
            else:
                if "item" not in dean_cats:
                    dean_cats["item"] = []
                if val not in dean_cats["item"]:
                    dean_cats["item"].append(val)
            dean_setting.value = json.dumps(dean_cats)
        else:
            if cat_type in ("source_of_fund", "expenditure"):
                dean_cats = {"source_of_fund": [val], "item": []}
            else:
                dean_cats = {"source_of_fund": [], "item": [val]}
            db.add(Settings(key_name="budget_categories_added_by_dean", value=json.dumps(dean_cats)))

    await db.commit()
    return await get_stored_categories(db)


@router.delete("/budget/categories")
async def delete_budget_category(
    type: str = Query(...),
    value: str = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    await db.refresh(current_user, ["role"])
    if current_user.role.group_key != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete categories")

    value = value.strip()
    if not value:
        raise HTTPException(status_code=400, detail="Category value is required")

    if type in ("source_of_fund", "expenditure"):
        key = "budget_source_of_fund_categories"
    elif type == "item":
        key = "budget_item_categories"
    else:
        raise HTTPException(status_code=400, detail="Invalid type: must be 'source_of_fund' or 'item'")

    # Load dean_setting to remove from there if needed
    result_dean = await db.execute(select(Settings).where(Settings.key_name == "budget_categories_added_by_dean"))
    dean_setting = result_dean.scalar_one_or_none()
    dean_cats = {"source_of_fund": [], "item": []}
    if dean_setting:
        import json
        try:
            dean_cats = json.loads(dean_setting.value)
        except Exception:
            dean_cats = {"source_of_fund": [], "item": []}

    # Verify if category exists in Settings
    result_setting = await db.execute(select(Settings).where(Settings.key_name == key))
    setting = result_setting.scalar_one_or_none()
    existing = []
    if setting:
        existing = [c.strip() for c in setting.value.split(",") if c.strip()]
    else:
        # Fallback default values
        if type in ("source_of_fund", "expenditure"):
            existing = ["CAPEX", "OPEX"]
        else:
            existing = ["computer", "lab_equipment", "software", "furniture"]

    if value not in existing:
        raise HTTPException(status_code=400, detail=f"Category '{value}' does not exist")

    # Check if category is currently used by any BudgetMaster entry
    if type in ("source_of_fund", "expenditure"):
        count_stmt = select(func.count(BudgetMaster.id)).where(BudgetMaster.source_of_fund == value)
    else:
        count_stmt = select(func.count(BudgetMaster.id)).where(BudgetMaster.category == value)

    count_res = await db.execute(count_stmt)
    use_count = count_res.scalar() or 0
    if use_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Category '{value}' is in use by {use_count} budget file(s) and cannot be deleted"
        )

    # Perform deletion from Settings
    if setting:
        existing.remove(value)
        setting.value = ",".join(existing)
    else:
        existing.remove(value)
        setting = Settings(key_name=key, value=",".join(existing))
        db.add(setting)

    # Remove from dean_cats metadata list if it exists there
    if type in ("source_of_fund", "expenditure"):
        if value in dean_cats.get("source_of_fund", []):
            dean_cats["source_of_fund"].remove(value)
    else:
        if value in dean_cats.get("item", []):
            dean_cats["item"].remove(value)

    if dean_setting:
        dean_setting.value = json.dumps(dean_cats)

    await db.commit()
    return await get_stored_categories(db)


# ─────────────────────────────────────────────────────────────────────────────
# SOURCE OF FUNDS MASTER (Configurable)
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/source-of-funds")
async def list_source_of_funds(db: AsyncSession = Depends(get_db), _=AdminDep):
    """Return all source of funds (active and inactive)."""
    result = await db.execute(select(SourceOfFund).order_by(SourceOfFund.name))
    funds = result.scalars().all()
    return [
        {
            "id": f.id,
            "name": f.name,
            "description": f.description,
            "is_active": f.is_active,
            "created_at": f.created_at.isoformat() if f.created_at else None,
        }
        for f in funds
    ]


@router.post("/source-of-funds")
async def create_source_of_fund(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    """Create a new source of fund."""
    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="name is required")

    existing = await db.execute(select(SourceOfFund).where(SourceOfFund.name == name))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail=f"Source of fund '{name}' already exists")

    sof = SourceOfFund(
        name=name,
        description=(body.get("description") or "").strip() or None,
        is_active=body.get("is_active", True),
    )
    db.add(sof)
    await db.commit()
    return {"message": "Source of fund created", "id": sof.id}


@router.put("/source-of-funds/{sof_id}")
async def update_source_of_fund(sof_id: int, body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    """Edit a source of fund (name, description, is_active)."""
    result = await db.execute(select(SourceOfFund).where(SourceOfFund.id == sof_id))
    sof = result.scalar_one_or_none()
    if not sof:
        raise HTTPException(status_code=404, detail="Source of fund not found")

    if "name" in body:
        new_name = (body["name"] or "").strip()
        if not new_name:
            raise HTTPException(status_code=400, detail="name cannot be empty")
        # Check uniqueness (excluding self)
        dup = await db.execute(
            select(SourceOfFund).where(SourceOfFund.name == new_name, SourceOfFund.id != sof_id)
        )
        if dup.scalar_one_or_none():
            raise HTTPException(status_code=409, detail=f"Name '{new_name}' is already used by another fund")
        sof.name = new_name
    if "description" in body:
        sof.description = (body["description"] or "").strip() or None
    if "is_active" in body:
        sof.is_active = body["is_active"]

    await db.commit()
    return {"message": "Source of fund updated"}


@router.patch("/source-of-funds/{sof_id}/toggle")
async def toggle_source_of_fund(sof_id: int, db: AsyncSession = Depends(get_db), _=AdminDep):
    """Toggle is_active for a source of fund."""
    result = await db.execute(select(SourceOfFund).where(SourceOfFund.id == sof_id))
    sof = result.scalar_one_or_none()
    if not sof:
        raise HTTPException(status_code=404, detail="Source of fund not found")
    sof.is_active = not sof.is_active
    await db.commit()
    return {"message": "Toggled", "is_active": sof.is_active}


@router.delete("/source-of-funds/{sof_id}")
async def delete_source_of_fund(sof_id: int, db: AsyncSession = Depends(get_db), _=AdminDep):
    """
    Hard-delete a source of fund only if it's not referenced anywhere.
    If it IS referenced, return 400 suggesting deactivation instead.
    """
    result = await db.execute(select(SourceOfFund).where(SourceOfFund.id == sof_id))
    sof = result.scalar_one_or_none()
    if not sof:
        raise HTTPException(status_code=404, detail="Source of fund not found")

    # Check references in BudgetMaster (by string name match for backward compat)
    bm_count_res = await db.execute(
        select(func.count(BudgetMaster.id)).where(BudgetMaster.source_of_fund == sof.name)
    )
    bm_count = bm_count_res.scalar() or 0

    # Check references in WorkFlowHierarchy (by FK)
    wf_count_res = await db.execute(
        select(func.count(WorkFlowHierarchy.id)).where(WorkFlowHierarchy.source_of_fund_id == sof_id)
    )
    wf_count = wf_count_res.scalar() or 0

    # Check references in AdministrativeApprovalWorkflow (by FK)
    from app.models.administrative_approval import AdministrativeApprovalWorkflow
    aaw_count_res = await db.execute(
        select(func.count(AdministrativeApprovalWorkflow.id)).where(AdministrativeApprovalWorkflow.source_of_fund_id == sof_id)
    )
    aaw_count = aaw_count_res.scalar() or 0

    total_refs = bm_count + wf_count + aaw_count
    if total_refs > 0:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Cannot delete '{sof.name}': it is referenced by "
                f"{bm_count} budget file(s), {wf_count} PR workflow step(s), and "
                f"{aaw_count} AA workflow step(s). Please deactivate it instead."
            )
        )

    await db.delete(sof)
    await db.commit()
    return {"message": f"Source of fund '{sof.name}' deleted"}


@router.get("/budget/next-file-number")

async def get_next_file_number(
    department_id: int,
    source_of_fund: str = Query(None),
    expenditure_category: str = Query(None),
    financial_year_id: int = Query(...),
    is_temporary: bool = Query(False),
    db: AsyncSession = Depends(get_db),
    _=BudgetViewDep
):
    from app.models.user import Department
    from app.models.budget import FinancialYear, BudgetMaster
    from sqlalchemy import func, or_

    fund = source_of_fund or expenditure_category
    if not fund:
        raise HTTPException(status_code=400, detail="source_of_fund or expenditure_category is required")

    dept_res = await db.execute(select(Department).where(Department.id == department_id))
    dept = dept_res.scalar_one_or_none()
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")

    fy_res = await db.execute(select(FinancialYear).where(FinancialYear.id == financial_year_id))
    fy = fy_res.scalar_one_or_none()
    if not fy:
        raise HTTPException(status_code=404, detail="Financial Year not found")

    is_temp = is_temporary if isinstance(is_temporary, bool) else False

    if is_temp:
        stmt = select(func.count(BudgetMaster.id)).where(
            and_(
                BudgetMaster.department_id == department_id,
                BudgetMaster.source_of_fund == fund,
                BudgetMaster.financial_year_id == financial_year_id,
                BudgetMaster.file_no.ilike("TEMP/%")
            )
        )
    else:
        stmt = select(func.count(BudgetMaster.id)).where(
            and_(
                BudgetMaster.department_id == department_id,
                BudgetMaster.source_of_fund == fund,
                BudgetMaster.financial_year_id == financial_year_id,
                or_(
                    BudgetMaster.file_no.ilike("NITT/%"),
                    ~BudgetMaster.file_no.ilike("TEMP/%")
                )
            )
        )
    count_res = await db.execute(stmt)
    count = count_res.scalar() or 0
    next_num = count + 1

    dept_code = dept.short_code.upper()
    source_code = fund.upper()
    fy_label = fy.label.upper()

    if is_temp:
        file_no = f"TEMP/F.No.{next_num:04d}/{source_code}/{fy_label}/{dept_code}"
    else:
        file_no = f"NITT/F.No.{next_num:04d}/{source_code}/{fy_label}/{dept_code}"
    return {"file_no": file_no}


async def generate_permanent_file_number(
    db: AsyncSession,
    department_id: int,
    source_of_fund: str,
    financial_year_id: int
) -> str:
    from app.models.user import Department
    from app.models.budget import FinancialYear, BudgetMaster
    from sqlalchemy import func, and_, or_

    dept_res = await db.execute(select(Department).where(Department.id == department_id))
    dept = dept_res.scalar_one_or_none()
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")

    fy_res = await db.execute(select(FinancialYear).where(FinancialYear.id == financial_year_id))
    fy = fy_res.scalar_one_or_none()
    if not fy:
        raise HTTPException(status_code=404, detail="Financial Year not found")

    stmt = select(func.count(BudgetMaster.id)).where(
        and_(
            BudgetMaster.department_id == department_id,
            BudgetMaster.source_of_fund == source_of_fund,
            BudgetMaster.financial_year_id == financial_year_id,
            or_(
                BudgetMaster.file_no.ilike("NITT/%"),
                ~BudgetMaster.file_no.ilike("TEMP/%")
            )
        )
    )
    count_res = await db.execute(stmt)
    count = count_res.scalar() or 0
    next_num = count + 1

    dept_code = dept.short_code.upper()
    source_code = source_of_fund.upper()
    fy_label = fy.label.upper()

    return f"NITT/F.No.{next_num:04d}/{source_code}/{fy_label}/{dept_code}"


@router.get("/budget/{b_id}")
async def get_budget_detail(b_id: int, db: AsyncSession = Depends(get_db), _=BudgetViewDep):
    result = await db.execute(
        select(BudgetMaster).where(BudgetMaster.id == b_id)
    )
    b = result.scalar_one_or_none()
    if not b:
        raise HTTPException(status_code=404, detail="Budget not found")
    committed = b.committed_amount if b.committed_amount is not None else 0.0
    utilized = b.utilized_amount if b.utilized_amount is not None else 0.0
    return {
        "id": b.id,
        "department_id": b.department_id,
        "financial_year_id": b.financial_year_id,
        "source_of_fund": b.source_of_fund,
        "expenditure_category": b.source_of_fund,
        "item_name": b.item_name,
        "category": b.category,
        "unit_cost": b.unit_cost,
        "quantity": b.quantity,
        "total_cost": b.total_allocation,
        "total_allocation": b.total_allocation,
        "committed_amount": committed,
        "locked_amount": committed,
        "utilized_amount": utilized,
        "deducted_amount": utilized,
        "available_balance": b.total_allocation - committed - utilized,
        "available_amount": b.total_allocation - committed - utilized,
        "file_no": b.file_no,
        "remarks": b.remarks,
        "expert1_id": b.expert1_id,
        "expert2_id": b.expert2_id,
        "director_faculty_id": b.director_faculty_id,
        "allocated_initiator_id": b.allocated_initiator_id,
        "attachment_path": b.attachment_path,
        "attachment_url": f"/static/uploads/{b.attachment_path}" if b.attachment_path else None,
        "project_code": b.project_code,
        "principal_investigator": b.principal_investigator,
        "project_due_date": b.project_due_date.isoformat() if b.project_due_date else None,
        "created_at": b.created_at.isoformat() if b.created_at else None,
    }


@router.post("/budget")
async def create_budget(
    department_id: Any = Form(...),
    financial_year_id: Any = Form(None),
    source_of_fund: Any = Form(None),
    item_name: Any = Form(None),
    category: Any = Form(None),
    unit_cost: Any = Form(None),
    quantity: Any = Form(None),
    file_no: Any = Form(None),
    remarks: Optional[str] = Form(None),
    course_code: str = Form("N/A"),
    attachment: Optional[UploadFile] = File(None),
    project_code: Optional[str] = Form(None),
    principal_investigator: Optional[str] = Form(None),
    project_due_date: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _ = None,
):
    """Create a budget file. Requires a supporting document attachment (PDF/image)."""
    is_dict_call = isinstance(department_id, dict)

    # Normalize attachment if it's the FastAPI parameter default
    if attachment is not None and not (isinstance(attachment, UploadFile) or hasattr(attachment, "file")):
        attachment = None

    if is_dict_call:
        body = department_id
        db_session = financial_year_id
        current_user = _ if _ is not None else user
        
        department_id = body.get("department_id")
        financial_year_id = body.get("financial_year_id")
        source_of_fund = body.get("source_of_fund")
        item_name = body.get("item_name")
        category = body.get("category")
        unit_cost = float(body.get("unit_cost", 0.0)) if body.get("unit_cost") is not None else 0.0
        quantity = int(body.get("quantity", 1)) if body.get("quantity") is not None else 1
        file_no = body.get("file_no")
        remarks = body.get("remarks")
        course_code = body.get("course_code", "N/A")
        project_code = body.get("project_code")
        principal_investigator = body.get("principal_investigator")
        project_due_date = body.get("project_due_date")
        if db_session is not None and not isinstance(db_session, (int, str)) and hasattr(db_session, "execute"):
            db = db_session
        user = current_user
    else:
        if department_id is not None:
            department_id = int(department_id)
        if financial_year_id is not None:
            financial_year_id = int(financial_year_id)
        if unit_cost is not None:
            unit_cost = float(unit_cost)
        else:
            unit_cost = 0.0
        if quantity is not None:
            quantity = int(quantity)
        else:
            quantity = 1

    if _ is not None and not is_dict_call:
        user = _

    await db.refresh(user, ["role"])
    group_key = user.role.group_key if user.role else None

    if group_key not in ("dean_approver", "admin", "hod"):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    # Allow Dean P&D (Budget) to create budgets

    if group_key == "hod" and user.department_id != department_id:
        raise HTTPException(status_code=403, detail="HOD can only create budget files for their own department")

    file_no_upper = file_no.upper() if file_no else ""
    if group_key == "hod" and not file_no_upper.startswith("TEMP/"):
        raise HTTPException(status_code=403, detail="HOD can only create temporary budget files (starting with TEMP/)")

    fy_res = await db.execute(select(FinancialYear).where(FinancialYear.id == financial_year_id))
    fy = fy_res.scalar_one_or_none()
    if not fy:
        raise HTTPException(status_code=400, detail="Financial Year not found")
    if fy.is_closed:
        raise HTTPException(status_code=400, detail="The selected financial year is closed.")

    # Validate and save the attachment
    rel_path = None
    if attachment is not None:
        import os, uuid as _uuid
        from app.core.config import settings as _settings

        ext = os.path.splitext(attachment.filename or "")[1].lower()
        if ext not in {".pdf", ".png", ".jpg", ".jpeg"}:
            raise HTTPException(status_code=400, detail="Attachment must be PDF, PNG, JPG, or JPEG.")

        content = await attachment.read()
        if len(content) > 10 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="Attachment size must be under 10 MB.")

        # Magic bytes validation
        header = content[:4]
        if ext == ".pdf" and not header.startswith(b"%PDF"):
            raise HTTPException(status_code=400, detail="Invalid PDF file.")
        elif ext == ".png" and not header.startswith(b"\x89PNG"):
            raise HTTPException(status_code=400, detail="Invalid PNG file.")
        elif ext in (".jpg", ".jpeg") and not header.startswith(b"\xff\xd8\xff"):
            raise HTTPException(status_code=400, detail="Invalid JPEG file.")

        rel_folder = os.path.join("budget_attachments")
        abs_folder = os.path.join(_settings.STORAGE_PATH, rel_folder)
        os.makedirs(abs_folder, exist_ok=True)
        filename = f"{_uuid.uuid4().hex}{ext}"
        rel_path = os.path.join(rel_folder, filename)
        abs_path = os.path.join(_settings.STORAGE_PATH, rel_path)
        with open(abs_path, "wb") as fh:
            fh.write(content)
    else:
        if not is_dict_call:
            raise HTTPException(status_code=400, detail="Supporting document attachment is compulsory.")

    if not is_dict_call:
        if not remarks or not remarks.strip():
            raise HTTPException(status_code=400, detail="Remarks / Justification is required")

    parsed_due_date = None
    if source_of_fund == "R&C":
        if not project_code or not project_code.strip():
            raise HTTPException(status_code=400, detail="Project Code is required for R&C source of fund")
        if not principal_investigator or not principal_investigator.strip():
            raise HTTPException(status_code=400, detail="Principal Investigator is required for R&C source of fund")
        if not project_due_date or not project_due_date.strip():
            raise HTTPException(status_code=400, detail="Project Due Date is required for R&C source of fund")
        
        try:
            parsed_due_date = datetime.strptime(project_due_date.strip(), "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format for Project Due Date. Must be YYYY-MM-DD")

    b = BudgetMaster(
        department_id=department_id,
        financial_year_id=financial_year_id,
        source_of_fund=source_of_fund,
        item_name=item_name,
        category=category,
        course_code=course_code,
        unit_cost=unit_cost,
        quantity=quantity,
        total_cost=round(unit_cost * quantity, 2),
        file_no=file_no_upper,
        remarks=remarks,
        is_revision=False,
        attachment_path=rel_path,
        project_code=project_code.strip() if project_code else None,
        principal_investigator=principal_investigator.strip() if principal_investigator else None,
        project_due_date=parsed_due_date,
    )
    db.add(b)
    await db.commit()
    return {"message": "Budget created", "id": b.id}


@router.put("/budget/{b_id}")
async def update_budget(b_id: int, body: dict, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user), _=None):
    if _ is not None:
        user = _
    await db.refresh(user, ["role"])
    group_key = user.role.group_key if user.role else None
    
    if group_key not in ("dean_approver", "admin", "hod"):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    if user.designation == "Dean P&D (Budget)":
        raise HTTPException(status_code=403, detail="Dean P&D (Budget) is not authorized to create/modify budgets")

    result = await db.execute(select(BudgetMaster).where(BudgetMaster.id == b_id))
    b = result.scalar_one_or_none()
    if not b:
        raise HTTPException(status_code=404, detail="Not found")

    if group_key == "hod" and b.department_id != user.department_id:
        raise HTTPException(status_code=403, detail="HOD can only update budget files for their own department")
        
    if group_key == "hod" and "department_id" in body and int(body["department_id"]) != user.department_id:
        raise HTTPException(status_code=403, detail="HOD cannot change department of the budget file")

    if group_key == "hod" and not b.file_no.upper().startswith("TEMP/"):
        raise HTTPException(status_code=403, detail="HOD cannot modify allocated permanent budget files")

    if group_key == "hod" and "file_no" in body:
        new_file_no = body["file_no"].upper()
        if not new_file_no.startswith("TEMP/"):
            raise HTTPException(status_code=403, detail="HOD can only assign temporary file numbers (starting with TEMP/)")

    # Check if current budget file belongs to closed financial year
    current_fy_res = await db.execute(select(FinancialYear).where(FinancialYear.id == b.financial_year_id))
    current_fy = current_fy_res.scalar_one_or_none()
    if current_fy and current_fy.is_closed:
        raise HTTPException(status_code=400, detail="The current financial year for this budget is closed. Budgets in closed financial years cannot be modified.")

    old_file_no = b.file_no
    b.item_name = body.get("item_name", b.item_name)
    if "department_id" in body:
        b.department_id = int(body["department_id"])
    if "financial_year_id" in body:
        new_fy_id = int(body["financial_year_id"])
        new_fy_res = await db.execute(select(FinancialYear).where(FinancialYear.id == new_fy_id))
        new_fy = new_fy_res.scalar_one_or_none()
        if not new_fy:
            raise HTTPException(status_code=400, detail="Financial Year not found")
        if new_fy.is_closed:
            raise HTTPException(status_code=400, detail="The target financial year is closed. Budgets in closed financial years cannot be modified.")
        b.financial_year_id = new_fy_id
    if "source_of_fund" in body:
        b.source_of_fund = body["source_of_fund"]
    elif "expenditure_category" in body:
        b.source_of_fund = body["expenditure_category"]
    
    if "project_code" in body:
        b.project_code = body["project_code"]
    if "principal_investigator" in body:
        b.principal_investigator = body["principal_investigator"]
    if "project_due_date" in body:
        val = body["project_due_date"]
        if val:
            try:
                b.project_due_date = datetime.strptime(str(val).strip(), "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date format for Project Due Date. Must be YYYY-MM-DD")
        else:
            b.project_due_date = None

    if b.source_of_fund == "R&C":
        if not b.project_code or not b.project_code.strip():
            raise HTTPException(status_code=400, detail="Project Code is required for R&C source of fund")
        if not b.principal_investigator or not b.principal_investigator.strip():
            raise HTTPException(status_code=400, detail="Principal Investigator is required for R&C source of fund")
        if not b.project_due_date:
            raise HTTPException(status_code=400, detail="Project Due Date is required for R&C source of fund")

    if "category" in body:
        b.category = body["category"]
    if "file_no" in body:
        new_file_no = body["file_no"].upper()
        if old_file_no.upper().startswith("TEMP") and not new_file_no.upper().startswith("TEMP"):
            remarks = body.get("remarks")
            if not remarks or not remarks.strip():
                raise HTTPException(
                    status_code=400,
                    detail="Remarks / justification is required for budget file number allocation",
                )
            new_file_no = await generate_permanent_file_number(
                db, b.department_id, b.source_of_fund, b.financial_year_id
            )
        b.file_no = new_file_no
    if "unit_cost" in body and "quantity" in body:
        b.unit_cost = float(body["unit_cost"])
        b.quantity = int(body["quantity"])
        b.total_cost = round(b.unit_cost * b.quantity, 2)
    if "allocated_initiator_id" in body:
        b.allocated_initiator_id = body["allocated_initiator_id"]
    if "remarks" in body:
        b.remarks = body["remarks"]

    # Check for TEMP -> non-TEMP transition and resume associated PRs
    if "file_no" in body:
        new_file_no = b.file_no
        if old_file_no.upper().startswith("TEMP") and not new_file_no.upper().startswith("TEMP"):
            stmt = (
                select(PurchaseRequest)
                .options(
                    selectinload(PurchaseRequest.flow),
                    selectinload(PurchaseRequest.items)
                )
                .join(PurchaseRequestItem)
                .where(
                    and_(
                        PurchaseRequest.current_status == RequestStatus.BUDGET_FILE_ALLOCATION,
                        PurchaseRequestItem.budget_file_id == b.id
                    )
                )
            )
            prs_res = await db.execute(stmt)
            prs_to_resume = prs_res.scalars().all()
            
            for pr in prs_to_resume:
                other_items_stmt = select(PurchaseRequestItem).where(
                    and_(
                        PurchaseRequestItem.purchase_request_id == pr.id,
                        PurchaseRequestItem.budget_file_id != b.id
                    )
                )
                other_items_res = await db.execute(other_items_stmt)
                other_items = other_items_res.scalars().all()
                
                all_clear = True
                for item in other_items:
                    await db.refresh(item, ["budget_file"])
                    if item.budget_file and item.budget_file.file_no.upper().startswith("TEMP"):
                        all_clear = False
                        break
                        
                if all_clear:
                    pr.current_status = RequestStatus.IN_PROGRESS
                    allocation_date_str = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
                    update_remarks = body.get("remarks", "")
                    history_remarks = f"Budget File Number: {new_file_no}\nAllocation Date: {allocation_date_str}\nRemarks: {update_remarks}"
                    
                    from app.services.flow_engine import FlowEngineService
                    flow_engine = FlowEngineService(db)
                    await flow_engine._add_history(
                        pr,
                        user,
                        "Budget File Allocated",
                        history_remarks
                    )
                    
                    # Notify next step
                    if pr.flow:
                        from app.models.purchase_request import WorkFlowHierarchy
                        from app.services.email_service import EmailService
                        new_step_result = await db.execute(
                            select(WorkFlowHierarchy).options(
                                selectinload(WorkFlowHierarchy.role),
                                selectinload(WorkFlowHierarchy.user)
                            ).where(
                                flow_engine._wf_filters(pr, pr.flow.phase_id, step_order=pr.flow.step_order)
                            )
                        )
                        new_step = new_step_result.scalar_one_or_none()
                        if new_step:
                            email_svc = EmailService(None)
                            if new_step.user_type == "user" and new_step.user_id:
                                user_res = await db.execute(select(User.email).where(User.id == new_step.user_id))
                                email = user_res.scalar_one_or_none()
                                next_emails = [email] if email else []
                            else:
                                next_emails = await flow_engine.get_next_approvers_emails(pr, new_step.user_group)
                            for email in next_emails:
                                email_svc.notify_next_approver(pr.id, pr.icr_number, new_step.role.name if new_step.role else new_step.user_group, email)

    await db.commit()
    return {"message": "Budget updated"}


@router.delete("/budget/clear")
async def clear_all_budgets(db: AsyncSession = Depends(get_db), _=DeanBudgetDep):
    """Deletes all budget master entries to start fresh. Skips any budgets linked to active PR items or belonging to closed financial years."""
    from app.models.purchase_request import PurchaseRequestItem
    from sqlalchemy import delete

    # Get linked budget IDs
    linked_res = await db.execute(select(PurchaseRequestItem.budget_file_id))
    linked_ids = {row[0] for row in linked_res.fetchall() if row[0] is not None}

    closed_fy_res = await db.execute(select(FinancialYear.id).where(FinancialYear.is_closed == True))
    closed_fy_ids = [row[0] for row in closed_fy_res.fetchall()]

    stmt = delete(BudgetMaster)
    if closed_fy_ids:
        stmt = stmt.where(BudgetMaster.financial_year_id.not_in(closed_fy_ids))
    if linked_ids:
        stmt = stmt.where(BudgetMaster.id.not_in(linked_ids))

    result = await db.execute(stmt)
    await db.commit()
    return {"message": f"Cleared {result.rowcount} unlinked budget files. Budgets in closed financial years and those linked to active Purchase Requests could not be deleted."}


# ─────────────────────────────────────────────────────────────────────────────
# PROCUREMENT METHODS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/procurement-methods")
async def list_procurement_methods(db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(ProcurementManager).order_by(ProcurementManager.name))
    return [
        {
            "id": p.id,
            "name": p.name,
            "description": p.description,
            "max_amount": p.max_amount,
            "form_schema": p.form_schema
        }
        for p in result.scalars()
    ]


@router.post("/procurement-methods")
async def create_procurement_method(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    p = ProcurementManager(
        name=body["name"],
        description=body.get("description"),
        max_amount=float(body["max_amount"]) if body.get("max_amount") is not None else None,
        form_schema=body.get("form_schema")
    )
    db.add(p)
    await db.commit()
    return {"message": "Procurement method created", "id": p.id}


@router.put("/procurement-methods/{pm_id}")
async def update_procurement_method(pm_id: int, body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(ProcurementManager).where(ProcurementManager.id == pm_id))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Procurement method not found")
    if "name" in body:
        p.name = body["name"]
    if "description" in body:
        p.description = body["description"]
    if "max_amount" in body:
        p.max_amount = float(body["max_amount"]) if body["max_amount"] is not None else None
    if "form_schema" in body:
        p.form_schema = body["form_schema"]
    await db.commit()
    return {"message": "Procurement method updated"}


@router.delete("/procurement-methods/{pm_id}")
async def delete_procurement_method(pm_id: int, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(ProcurementManager).where(ProcurementManager.id == pm_id))
    p = result.scalar_one_or_none()
    if p:
        await db.delete(p)
        await db.commit()
    return {"message": "Procurement method deleted"}


# ─────────────────────────────────────────────────────────────────────────────
# WORKFLOWS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/workflows")
async def list_workflows(db: AsyncSession = Depends(get_db), _=AdminDep):
    from sqlalchemy.orm import selectinload
    result = await db.execute(
        select(WorkFlowHierarchy).options(
            selectinload(WorkFlowHierarchy.role),
            selectinload(WorkFlowHierarchy.user)
        ).order_by(
            WorkFlowHierarchy.category_id,
            WorkFlowHierarchy.procurement_id,
            WorkFlowHierarchy.phase_id,
            WorkFlowHierarchy.step_order,
        )
    )
    entries = result.scalars().all()
    return [
        {
            "id": w.id,
            "category_id": w.category_id,
            "procurement_id": w.procurement_id,
            "phase_id": w.phase_id,
            "step_order": w.step_order,
            "user_type": w.user_type,
            "user_id": w.user_id,
            "user_name": w.user.name if w.user else None,
            "user_group": w.user_group,
            "role_id": w.role_id,
            "role_name": w.role.name if w.role else (w.user_group.replace("_", " ").title() if w.user_group else None),
            "is_enabled": w.is_enabled,
            "purchase_type": w.purchase_type,
            "tender_vendors_threshold": w.tender_vendors_threshold,
            "tender_vendors_comparison": w.tender_vendors_comparison,
            "skip_condition": w.skip_condition,
            "condition_field": w.condition_field,
            "condition_operator": w.condition_operator,
            "condition_value": w.condition_value,
            "source_of_fund_id": w.source_of_fund_id,
            "committee_size": w.committee_size,
        }
        for w in entries
    ]


@router.post("/workflows")
async def create_workflow(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    role_id = body.get("role_id")
    user_id = body.get("user_id")
    user_group = body.get("user_group")
    user_type = body.get("user_type", "verifier")

    if user_type == "user" and user_id:
        user_res = await db.execute(select(User).where(User.id == user_id))
        user_obj = user_res.scalar_one_or_none()
        if not user_obj:
            raise HTTPException(status_code=400, detail="User not found")
        role_id = None
        user_group = None
    elif user_type in ["purchase_initiator", "da_assigner", "verifier_da", "tech_evaluation"]:
        user_id = None
        role_id = None
        user_group = None
    else:
        if user_type not in ["verifier", "approver", "partial_approver"]:
            user_type = "verifier"
        user_id = None
        if role_id:
            role_res = await db.execute(select(RoleManager).where(RoleManager.id == role_id))
            role_obj = role_res.scalar_one_or_none()
            if role_obj:
                user_group = role_obj.group_key
                
    purchase_type = body.get("purchase_type", "research")
    wf = WorkFlowHierarchy(
        category_id=body["category_id"],
        phase_id=body["phase_id"],
        procurement_id=body["procurement_id"],
        step_order=body["step_order"],
        user_type=user_type,
        user_id=user_id,
        user_group=user_group,
        role_id=role_id,
        purchase_type=purchase_type,
        is_enabled=True,
        tender_vendors_threshold=body.get("tender_vendors_threshold"),
        tender_vendors_comparison=body.get("tender_vendors_comparison"),
        skip_condition=body.get("skip_condition"),
        condition_field=body.get("condition_field"),
        condition_operator=body.get("condition_operator"),
        condition_value=body.get("condition_value"),
        source_of_fund_id=body.get("source_of_fund_id"),  # None = any fund (default)
        committee_size=body.get("committee_size"),  # Only for tech_evaluation steps
    )
    db.add(wf)
    await db.commit()
    return {"message": "Workflow created", "id": wf.id}


@router.put("/workflows/{wf_id}")
async def update_workflow(wf_id: int, body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(WorkFlowHierarchy).where(WorkFlowHierarchy.id == wf_id))
    wf = result.scalar_one_or_none()
    if not wf:
        raise HTTPException(status_code=404, detail="Not found")
    if "step_order" in body:
        wf.step_order = body["step_order"]
    if "user_type" in body:
        wf.user_type = body["user_type"]
        if wf.user_type in ["purchase_initiator", "da_assigner", "verifier_da", "tech_evaluation"]:
            wf.user_id = None
            wf.role_id = None
            wf.user_group = None
        elif wf.user_type == "user":
            wf.role_id = None
            wf.user_group = None
        elif wf.user_type not in ["verifier", "approver", "partial_approver"]:
            pass
    if "user_id" in body:
        wf.user_id = body["user_id"]
        if wf.user_id:
            wf.role_id = None
            wf.user_group = None
            wf.user_type = "user"
    if "role_id" in body:
        wf.role_id = body["role_id"]
        if wf.role_id:
            role_res = await db.execute(select(RoleManager).where(RoleManager.id == wf.role_id))
            role_obj = role_res.scalar_one_or_none()
            if role_obj:
                wf.user_group = role_obj.group_key
                wf.user_id = None
                if wf.user_type not in ["verifier", "approver", "partial_approver"]:
                    wf.user_type = "verifier"
        else:
            wf.role_id = None
    elif "user_group" in body:
        wf.user_group = body["user_group"]
        wf.user_id = None
        wf.role_id = None
        if wf.user_type not in ["verifier", "approver", "partial_approver"]:
            wf.user_type = "verifier"
    if "is_enabled" in body:
        wf.is_enabled = bool(body["is_enabled"])
    if "tender_vendors_threshold" in body:
        # Accept null/None to clear the threshold
        wf.tender_vendors_threshold = body["tender_vendors_threshold"]
    if "tender_vendors_comparison" in body:
        wf.tender_vendors_comparison = body["tender_vendors_comparison"]
    if "skip_condition" in body:
        wf.skip_condition = body["skip_condition"]
    if "condition_field" in body:
        wf.condition_field = body["condition_field"]
    if "condition_operator" in body:
        wf.condition_operator = body["condition_operator"]
    if "condition_value" in body:
        wf.condition_value = body["condition_value"]
    if "source_of_fund_id" in body:
        wf.source_of_fund_id = body["source_of_fund_id"]
    if "committee_size" in body:
        wf.committee_size = body["committee_size"]  # int 1/2/3 or null
    await db.commit()
    return {"message": "Workflow updated"}


@router.post("/workflows/reorder")
async def reorder_workflows(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    # Accept either step_ids (ordered list → assign 1,2,3...) or legacy steps [{id,step_order}]
    step_ids = body.get("step_ids", [])
    if step_ids:
        for idx, sid in enumerate(step_ids, start=1):
            res = await db.execute(select(WorkFlowHierarchy).where(WorkFlowHierarchy.id == sid))
            wf = res.scalar_one_or_none()
            if wf:
                wf.step_order = idx
    else:
        for item in body.get("steps", []):
            res = await db.execute(select(WorkFlowHierarchy).where(WorkFlowHierarchy.id == item["id"]))
            wf = res.scalar_one_or_none()
            if wf:
                wf.step_order = item["step_order"]
    await db.commit()
    return {"message": "Steps reordered"}


@router.patch("/workflows/{wf_id}/toggle")
async def toggle_workflow_step(wf_id: int, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(WorkFlowHierarchy).where(WorkFlowHierarchy.id == wf_id))
    wf = result.scalar_one_or_none()
    if not wf:
        raise HTTPException(status_code=404, detail="Not found")
    wf.is_enabled = not wf.is_enabled
    await db.commit()
    return {"message": "Toggled", "is_enabled": wf.is_enabled}


@router.delete("/workflows/{wf_id}")
async def delete_workflow(wf_id: int, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(WorkFlowHierarchy).where(WorkFlowHierarchy.id == wf_id))
    wf = result.scalar_one_or_none()
    if not wf:
        return {"message": "Workflow deleted"}

    # Safeguard: block delete if any active PR is currently at this exact step
    active_count_res = await db.execute(
        select(func.count(PurchaseRequestFlow.id))
        .join(PurchaseRequest, PurchaseRequest.id == PurchaseRequestFlow.purchase_request_id)
        .where(
            and_(
                PurchaseRequestFlow.phase_id == wf.phase_id,
                PurchaseRequestFlow.step_order == wf.step_order,
                PurchaseRequest.category_id == wf.category_id,
                PurchaseRequest.procurement_id == wf.procurement_id,
                PurchaseRequest.purchase_type == wf.purchase_type,
                PurchaseRequest.current_status.in_([
                    RequestStatus.PR_SUBMITTED,
                    RequestStatus.IN_PROGRESS,
                    RequestStatus.SENT_BACK,
                ])
            )
        )
    )
    active_count = active_count_res.scalar() or 0
    if active_count:
        raise HTTPException(
            status_code=409,
            detail=f"Cannot delete: {active_count} active purchase indent(s) are currently at this workflow step. Disable the step instead."
        )

    cat_id = wf.category_id
    proc_id = wf.procurement_id
    phase_id = wf.phase_id
    p_type = wf.purchase_type
    sof_id = wf.source_of_fund_id
    await db.delete(wf)
    await db.flush()
    # Renumber remaining steps in the same phase sequentially to close the gap
    remaining_res = await db.execute(
        select(WorkFlowHierarchy)
        .where(and_(
            WorkFlowHierarchy.category_id == cat_id,
            WorkFlowHierarchy.procurement_id == proc_id,
            WorkFlowHierarchy.phase_id == phase_id,
            WorkFlowHierarchy.purchase_type == p_type,
            WorkFlowHierarchy.source_of_fund_id == sof_id,
        ))
        .order_by(WorkFlowHierarchy.step_order)
    )
    remaining = remaining_res.scalars().all()
    for idx, rem in enumerate(remaining, start=1):
        rem.step_order = idx
    await db.commit()
    return {"message": "Workflow deleted"}


@router.post("/workflows/reset-defaults")
async def reset_workflows(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    """Reset workflow steps for one category × procurement × purchase type to seeded defaults."""
    from sqlalchemy import delete
    from app.seed_workflows import build_workflow_steps

    cat_id = body.get("category_id")
    proc_id = body.get("procurement_id")
    purchase_type = body.get("purchase_type", "research")
    # source_of_fund_id=None resets the default (any-fund) variant only
    source_of_fund_id = body.get("source_of_fund_id", None)
    if not cat_id or not proc_id:
        raise HTTPException(status_code=400, detail="Missing category_id or procurement_id")

    cat_res = await db.execute(select(PurchaseCategory).where(PurchaseCategory.id == cat_id))
    cat = cat_res.scalar_one_or_none()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")

    proc_res = await db.execute(select(ProcurementManager).where(ProcurementManager.id == proc_id))
    proc = proc_res.scalar_one_or_none()
    if not proc:
        raise HTTPException(status_code=404, detail="Procurement method not found")

    # Scope delete to the specific source_of_fund_id variant only
    delete_filter = and_(
        WorkFlowHierarchy.category_id == cat_id,
        WorkFlowHierarchy.procurement_id == proc_id,
        WorkFlowHierarchy.purchase_type == purchase_type,
        WorkFlowHierarchy.source_of_fund_id == source_of_fund_id,
    )
    await db.execute(delete(WorkFlowHierarchy).where(delete_filter))

    roles_res = await db.execute(select(RoleManager))
    roles = {r.value: r for r in roles_res.scalars()}

    phases_res = await db.execute(select(PhaseManager))
    phases = {}
    for p in phases_res.scalars():
        key = {
            "Administrative Approval": "AA",
            "Indent and Detailed Tech Specification": "AA",
            "Tendering": "TD",
            "Technical Evaluation": "TE",
            "Financial Sanction": "FS",
            "Purchase Order": "PO"
        }.get(p.phase_name)
        if key:
            phases[key] = p

    cat_key = "cat1" if cat.max_amount <= 100_000 else ("cat2" if cat.max_amount <= 1_000_000 else "cat3")
    categories = {cat_key: cat}

    all_rows = build_workflow_steps(roles, phases, categories, [proc])
    for w in all_rows:
        if w.purchase_type == purchase_type:
            w.source_of_fund_id = source_of_fund_id
            db.add(w)

    await db.commit()
    return {"message": "Workflows reset to defaults"}



# ─────────────────────────────────────────────────────────────────────────────
# PHASES & CATEGORIES
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/phases")
async def list_phases(db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(PhaseManager).order_by(PhaseManager.phase_order))
    return [{"id": p.id, "phase_name": p.phase_name} for p in result.scalars()]


@router.get("/categories")
async def list_categories(procurement_id: Optional[int] = None, db: AsyncSession = Depends(get_db), _=AdminDep):
    stmt = select(PurchaseCategory)
    if procurement_id is not None:
        stmt = stmt.where(PurchaseCategory.procurement_id == procurement_id)
    stmt = stmt.order_by(PurchaseCategory.procurement_id, PurchaseCategory.min_amount)
    result = await db.execute(stmt)
    return [
        {
            "id": c.id,
            "title": c.title,
            "min_amount": c.min_amount,
            "max_amount": c.max_amount,
            "is_active": c.is_active,
            "procurement_id": c.procurement_id,
            "requirement_type": c.requirement_type,
        }
        for c in result.scalars()
    ]


async def _sync_procurement_max_amount(db: AsyncSession, procurement_id: int) -> None:
    """Sync procurement_managers.max_amount to the highest category max_amount for that procurement.
    Only updates if the procurement method already has a non-null max_amount (i.e. it has a configured limit)."""
    pm_res = await db.execute(select(ProcurementManager).where(ProcurementManager.id == procurement_id))
    pm = pm_res.scalar_one_or_none()
    if pm is None or pm.max_amount is None:
        return
    max_res = await db.execute(
        select(func.max(PurchaseCategory.max_amount))
        .where(PurchaseCategory.procurement_id == procurement_id)
    )
    highest = max_res.scalar()
    if highest is not None and pm.max_amount != highest:
        pm.max_amount = highest


@router.post("/categories")
async def create_category(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    proc_id = int(body["procurement_id"])
    c = PurchaseCategory(
        title=body["title"],
        min_amount=float(body["min_amount"]),
        max_amount=float(body["max_amount"]),
        is_active=body.get("is_active", True),
        procurement_id=proc_id,
        requirement_type=body.get("requirement_type") if body.get("requirement_type") else None
    )
    db.add(c)
    await db.flush()
    await _sync_procurement_max_amount(db, proc_id)
    await db.commit()
    return {"message": "Category created", "id": c.id}


@router.put("/categories/{cat_id}")
async def update_category(cat_id: int, body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(PurchaseCategory).where(PurchaseCategory.id == cat_id))
    c = result.scalar_one_or_none()
    if not c:
        raise HTTPException(status_code=404, detail="Category not found")
    if "title" in body:
        c.title = body["title"]
    if "min_amount" in body:
        c.min_amount = float(body["min_amount"])
    if "max_amount" in body:
        c.max_amount = float(body["max_amount"])
    if "is_active" in body:
        c.is_active = bool(body["is_active"])
    if "procurement_id" in body:
        c.procurement_id = int(body["procurement_id"])
    if "requirement_type" in body:
        c.requirement_type = body["requirement_type"] if body["requirement_type"] else None
    await db.flush()
    await _sync_procurement_max_amount(db, c.procurement_id)
    await db.commit()
    return {"message": "Category updated"}


@router.delete("/categories/{cat_id}")
async def delete_category(cat_id: int, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(PurchaseCategory).where(PurchaseCategory.id == cat_id))
    c = result.scalar_one_or_none()
    if c:
        await db.delete(c)
        await db.commit()
    return {"message": "Category deleted"}


# ─────────────────────────────────────────────────────────────────────────────
# PENDING USER ONBOARDING
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/users/pending")
async def get_pending_users(db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(
        select(User)
        .options(selectinload(User.role), selectinload(User.department))
        .where(User.is_approved == False)
        .order_by(User.created_at.desc())
    )
    users = result.scalars().all()
    return [
        {
            "id": u.id,
            "title": u.title,
            "name": u.name,
            "email": u.email,
            "designation": u.designation,
            "gender": u.gender,
            "role": {"group_key": u.role.group_key, "name": u.role.name} if u.role else None,
            "department": {"name": u.department.name, "short_code": u.department.short_code} if u.department else None,
            "signature_path": f"/storage/{u.signature_path}" if u.signature_path else None,
            "created_at": u.created_at.isoformat() if u.created_at else None,
        }
        for u in users
    ]


@router.post("/users/{user_id}/approve")
async def approve_user(user_id: int, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(User).where(User.id == user_id))
    u = result.scalar_one_or_none()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    u.is_approved = True
    await db.commit()
    return {"message": "User approved successfully"}


@router.post("/users/{user_id}/reject")
async def reject_user(user_id: int, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(User).where(User.id == user_id))
    u = result.scalar_one_or_none()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    if u.signature_path:
        import os
        abs_path = os.path.join(settings.STORAGE_PATH, u.signature_path)
        if os.path.exists(abs_path):
            try:
                os.remove(abs_path)
            except Exception:
                pass
    await db.delete(u)
    await db.commit()
    return {"message": "User onboarding request rejected and deleted"}


# ─────────────────────────────────────────────────────────────────────────────
# BUDGET CSV EXPORT & IMPORT
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/budget/import-template")
async def download_budget_template(_=DeanBudgetDep):
    import io
    import csv
    from fastapi.responses import StreamingResponse

    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(["S.No", "Department", "File No", "Procurement", "Budget Amount (INR)"])
    writer.writerow(["1", "CSE", "NITT/CSE/2026-27/005", "GPU Server Purchase", "1500000"])
    writer.writerow(["2", "ECE", "NITT/ECE/2026-27/001", "Lab Oscilloscopes", "800000"])

    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="budget_import_template.csv"'}
    )


@router.post("/budget/import")
@limiter.limit("10/minute")
async def import_budget_csv(
    request: Request,
    file: UploadFile,
    financial_year_id: Optional[int] = Form(None),
    db: AsyncSession = Depends(get_db),
    _=DeanBudgetDep
):
    from app.services.import_service import ImportService
    content = await file.read()
    import_service = ImportService(db)
    return await import_service.import_budget_csv(content, file.filename, financial_year_id)


@router.patch("/users/{user_id}/role")
async def update_user_role(user_id: int, body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    role_id = body.get("role_id")
    if not role_id:
        raise HTTPException(status_code=400, detail="role_id is required")
        
    user_res = await db.execute(select(User).where(User.id == user_id))
    u = user_res.scalar_one_or_none()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
        
    role_res = await db.execute(select(RoleManager).where(RoleManager.id == role_id))
    role = role_res.scalar_one_or_none()
    if not role:
        raise HTTPException(status_code=400, detail="Invalid role selected")
        
    u.role_id = role_id
    await db.commit()
    return {"message": "User role updated successfully"}


@router.post("/purchase-requests/{pr_id}/force-advance")
async def force_advance_pr(
    pr_id: int,
    body: dict,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_roles("admin"))
):
    remarks = body.get("remarks")
    if not remarks or not remarks.strip():
        raise HTTPException(status_code=400, detail="Remarks are mandatory for force-advancing")
        
    result = await db.execute(select(PurchaseRequest).where(PurchaseRequest.id == pr_id))
    pr = result.scalar_one_or_none()
    if not pr:
        raise HTTPException(status_code=404, detail="Purchase request not found")
        
    from app.services.flow_engine import FlowEngineService
    flow_engine = FlowEngineService(db)
    
    try:
        await flow_engine.force_advance(pr, user, remarks)
        await db.commit()
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return {"message": "Purchase request force-advanced successfully"}


@router.post("/designations")
async def add_designation(
    body: dict,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user)
):
    await db.refresh(user, ["role"])
    if user.role.group_key != "admin":
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    value = body.get("value", "").strip()
    if not value:
        raise HTTPException(status_code=400, detail="Designation value is required")

    result = await db.execute(select(Settings).where(Settings.key_name == "designations"))
    setting = result.scalar_one_or_none()
    existing = []
    if setting:
        existing = [d.strip() for d in setting.value.split(",") if d.strip()]
        if value in existing:
            raise HTTPException(status_code=400, detail="Designation already exists")
        existing.append(value)
        setting.value = ",".join(existing)
    else:
        existing = ["Assistant Professor", "Associate Professor", "Professor", "Dean P&D (Budget)", "Dean P&D", "Director", "Registrar"]
        if value in existing:
            raise HTTPException(status_code=400, detail="Designation already exists")
        existing.append(value)
        setting = Settings(key_name="designations", value=",".join(existing))
        db.add(setting)

    await db.commit()
    return {"message": "Designation added successfully", "designations": existing}


@router.delete("/designations")
async def delete_designation(
    value: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user)
):
    await db.refresh(user, ["role"])
    if user.role.group_key != "admin":
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    value = value.strip()
    if not value:
        raise HTTPException(status_code=400, detail="Designation value is required")

    # Check if designation is currently used by any user
    user_check = await db.execute(select(User).where(User.designation == value))
    if user_check.scalars().first() is not None:
        raise HTTPException(status_code=400, detail="Designation is currently assigned to users and cannot be deleted")

    result = await db.execute(select(Settings).where(Settings.key_name == "designations"))
    setting = result.scalar_one_or_none()
    existing = []
    if setting:
        existing = [d.strip() for d in setting.value.split(",") if d.strip()]
    else:
        existing = ["Assistant Professor", "Associate Professor", "Professor", "Dean P&D (Budget)", "Dean P&D", "Director", "Registrar"]

    if value not in existing:
        raise HTTPException(status_code=400, detail="Designation does not exist")

    existing.remove(value)
    if setting:
        setting.value = ",".join(existing)
    else:
        setting = Settings(key_name="designations", value=",".join(existing))
        db.add(setting)

    await db.commit()
    return {"message": "Designation deleted successfully", "designations": existing}


# ─────────────────────────────────────────────────────────────────────────────
# ADMINISTRATIVE APPROVAL WORKFLOW
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/aa-workflows")
async def list_aa_workflows(db: AsyncSession = Depends(get_db), _=AdminDep):
    from app.models.administrative_approval import AdministrativeApprovalWorkflow
    result = await db.execute(
        select(AdministrativeApprovalWorkflow)
        .options(selectinload(AdministrativeApprovalWorkflow.role))
        .order_by(
            AdministrativeApprovalWorkflow.category_id,
            AdministrativeApprovalWorkflow.procurement_id,
            AdministrativeApprovalWorkflow.purchase_type,
            AdministrativeApprovalWorkflow.step_order,
        )
    )
    steps = result.scalars().all()
    return [
        {
            "id": s.id,
            "category_id": s.category_id,
            "procurement_id": s.procurement_id,
            "purchase_type": s.purchase_type,
            "step_order": s.step_order,
            "role_id": s.role_id,
            "role_name": s.role.name if s.role else None,
            "user_group": s.user_group,
            "is_enabled": s.is_enabled,
            "skip_condition": s.skip_condition,
            "source_of_fund_id": s.source_of_fund_id,
        }
        for s in steps
    ]


@router.post("/aa-workflows")
async def create_aa_workflow(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    from app.models.administrative_approval import AdministrativeApprovalWorkflow
    
    role_id = body.get("role_id")
    user_group = body.get("user_group")
    step_order = body.get("step_order")
    category_id = body.get("category_id")
    procurement_id = body.get("procurement_id")
    purchase_type = body.get("purchase_type")
    skip_condition = body.get("skip_condition")
    
    if not user_group:
        raise HTTPException(status_code=400, detail="user_group is required")
        
    if step_order is None:
        res = await db.execute(
            select(func.max(AdministrativeApprovalWorkflow.step_order))
            .where(
                and_(
                    AdministrativeApprovalWorkflow.category_id == category_id,
                    AdministrativeApprovalWorkflow.procurement_id == procurement_id,
                    AdministrativeApprovalWorkflow.purchase_type == purchase_type,
                )
            )
        )
        max_order = res.scalar() or 0
        step_order = max_order + 1
        
    s = AdministrativeApprovalWorkflow(
        role_id=role_id,
        user_group=user_group,
        step_order=step_order,
        is_enabled=True,
        category_id=category_id,
        procurement_id=procurement_id,
        purchase_type=purchase_type,
        skip_condition=skip_condition,
        source_of_fund_id=body.get("source_of_fund_id"),  # None = any fund (default)
    )
    db.add(s)
    await db.commit()
    return {"message": "Workflow step created", "id": s.id}


@router.put("/aa-workflows/{step_id}")
async def update_aa_workflow(step_id: int, body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    from app.models.administrative_approval import AdministrativeApprovalWorkflow
    result = await db.execute(select(AdministrativeApprovalWorkflow).where(AdministrativeApprovalWorkflow.id == step_id))
    s = result.scalar_one_or_none()
    if not s:
        raise HTTPException(status_code=404, detail="Step not found")
        
    if "role_id" in body:
        s.role_id = body["role_id"]
    if "user_group" in body:
        s.user_group = body["user_group"]
    if "step_order" in body:
        s.step_order = body["step_order"]
    if "is_enabled" in body:
        s.is_enabled = body["is_enabled"]
    if "category_id" in body:
        s.category_id = body["category_id"]
    if "procurement_id" in body:
        s.procurement_id = body["procurement_id"]
    if "purchase_type" in body:
        s.purchase_type = body["purchase_type"]
    if "skip_condition" in body:
        s.skip_condition = body["skip_condition"]
    if "source_of_fund_id" in body:
        s.source_of_fund_id = body["source_of_fund_id"]
        
    await db.commit()
    return {"message": "Workflow step updated"}


@router.delete("/aa-workflows/{step_id}")
async def delete_aa_workflow(step_id: int, db: AsyncSession = Depends(get_db), _=AdminDep):
    from app.models.administrative_approval import AdministrativeApprovalWorkflow, AdministrativeApproval
    result = await db.execute(select(AdministrativeApprovalWorkflow).where(AdministrativeApprovalWorkflow.id == step_id))
    s = result.scalar_one_or_none()
    if not s:
        raise HTTPException(status_code=404, detail="Step not found")

    # Safeguard: block delete if any active AA is currently pending at this step's user_group
    active_count_res = await db.execute(
        select(func.count(AdministrativeApproval.id))
        .where(AdministrativeApproval.pending_with == s.user_group)
    )
    active_count = active_count_res.scalar() or 0
    if active_count:
        raise HTTPException(
            status_code=409,
            detail=f"Cannot delete: {active_count} active administrative approval(s) are currently pending with '{s.user_group}'. Disable the step instead."
        )

    cat_id = s.category_id
    proc_id = s.procurement_id
    p_type = s.purchase_type
    sof_id = s.source_of_fund_id
    
    await db.delete(s)
    
    # Re-normalize step orders within same SoF variant
    remaining_res = await db.execute(
        select(AdministrativeApprovalWorkflow)
        .where(
            and_(
                AdministrativeApprovalWorkflow.category_id == cat_id,
                AdministrativeApprovalWorkflow.procurement_id == proc_id,
                AdministrativeApprovalWorkflow.purchase_type == p_type,
                AdministrativeApprovalWorkflow.source_of_fund_id == sof_id,
            )
        )
        .order_by(AdministrativeApprovalWorkflow.step_order)
    )
    remaining = remaining_res.scalars().all()
    for idx, rem in enumerate(remaining, start=1):
        rem.step_order = idx
        
    await db.commit()
    return {"message": "Workflow step deleted"}


@router.patch("/aa-workflows/{step_id}/toggle")
async def toggle_aa_workflow(step_id: int, db: AsyncSession = Depends(get_db), _=AdminDep):
    from app.models.administrative_approval import AdministrativeApprovalWorkflow
    result = await db.execute(select(AdministrativeApprovalWorkflow).where(AdministrativeApprovalWorkflow.id == step_id))
    s = result.scalar_one_or_none()
    if not s:
        raise HTTPException(status_code=404, detail="Step not found")
        
    s.is_enabled = not s.is_enabled
    await db.commit()
    return {"message": "Toggled", "is_enabled": s.is_enabled}


@router.post("/aa-workflows/reorder")
async def reorder_aa_workflows(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    from app.models.administrative_approval import AdministrativeApprovalWorkflow
    step_ids = body.get("step_ids", [])
    if not step_ids:
        raise HTTPException(status_code=400, detail="Missing step_ids")
        
    for idx, sid in enumerate(step_ids, start=1):
        await db.execute(
            text("UPDATE administrative_approval_workflows SET step_order = :order WHERE id = :id"),
            {"order": idx, "id": sid}
        )
    await db.commit()
    return {"message": "Reordered successfully"}


@router.post("/aa-workflows/reset-defaults")
async def reset_aa_workflows(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    from app.models.administrative_approval import AdministrativeApprovalWorkflow
    from app.models.user import RoleManager
    from sqlalchemy import delete

    category_id = body.get("category_id")  # None = global default
    procurement_id = body.get("procurement_id")  # None = global default
    purchase_type = body.get("purchase_type")  # None = global default
    source_of_fund_id = body.get("source_of_fund_id", None)

    is_global = category_id is None and procurement_id is None

    # Build delete condition matching exact null/non-null values
    conditions = [
        AdministrativeApprovalWorkflow.source_of_fund_id == source_of_fund_id,
    ]
    if is_global:
        conditions += [
            AdministrativeApprovalWorkflow.category_id == None,
            AdministrativeApprovalWorkflow.procurement_id == None,
            AdministrativeApprovalWorkflow.purchase_type == None,
        ]
    else:
        conditions += [
            AdministrativeApprovalWorkflow.category_id == category_id,
            AdministrativeApprovalWorkflow.procurement_id == procurement_id,
            AdministrativeApprovalWorkflow.purchase_type == purchase_type,
        ]

    await db.execute(delete(AdministrativeApprovalWorkflow).where(and_(*conditions)))

    if is_global:
        # Re-seed global defaults: NULL/NULL/NULL catch-all steps
        roles_res = await db.execute(select(RoleManager))
        roles = {r.value: r for r in roles_res.scalars()}
        dean_role = roles.get("dean_pd") or roles.get("dean")
        for step_order, group, role_key in [
            (1, "HOD", "hod"),
            (2, "ADPD", "adpd"),
            (3, "Dean", None),
        ]:
            role = roles.get(role_key) if role_key else dean_role
            db.add(AdministrativeApprovalWorkflow(
                category_id=None,
                procurement_id=None,
                purchase_type=None,
                step_order=step_order,
                user_group=group,
                role_id=role.id if role else None,
                source_of_fund_id=None,
            ))
        await db.commit()
        return {"message": "Global defaults reset to HOD → ADPD → Dean"}

    # Specific combination: just delete the override — global defaults now apply
    await db.commit()
    return {"message": "Override removed. Global default flow will now apply."}


